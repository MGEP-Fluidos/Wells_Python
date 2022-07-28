##########################################################################################################################################
##########################################################################################################################################
###################################################################TODO###################################################################
##########################################################################################################################################
##########################################################################################################################################

# instantiate_turbine(): TODO DOCUMENTATION
# fitness_func(): TODO DOCUMENTATION
# on_start(): TODO MODIFY DOCUMENTATION, METHOD'S ARG SPECIFICATION
# on_fitness(): TODO MODIFY DOCUMENTATION, METHOD'S ARG SPECIFICATION
# on_generation(): TODO MODIFY DOCUMENTATION, METHOD'S ARG SPECIFICATION
# on_stop(): TODO MODIFY DOCUMENTATION, METHOD'S ARG SPECIFICATION

##########################################################################################################################################
##########################################################################################################################################
############################################################IMPORTING PACKAGES############################################################
##########################################################################################################################################
##########################################################################################################################################

import numpy as np
import scipy as sc
import scipy.integrate as scint
import matplotlib as mpl
import matplotlib.pyplot as plt
import os
from scipy.interpolate import interp1d
from scipy.interpolate import LSQUnivariateSpline as lsqus
from scipy.integrate import simpson
from operator import itemgetter
import openpyxl
import pickle
import copy
from functools import partial
import Math_Tools.MathTools as mt
from scipy.interpolate import interp1d
from typing import Union
import itertools

##########################################################################################################################################
##########################################################################################################################################
#################################################TURBINE-CLASS-RELATED CLASS DEFINITIONS##################################################
##########################################################################################################################################
##########################################################################################################################################


#--------------------------------------------------------------------------------------------------------------------#
#---------------------------------------------------CLASS geom_params------------------------------------------------#             
#--------------------------------------------------------------------------------------------------------------------#
class geom_params:
    '''Geometrical parameters of a turbine stage.'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------__init__() METHOD------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def __init__(self,
                 ttype: str='Wells',
                 N: int=50,
                 rcas: float=0.25,
                 hub_to_tip_ratio: float=0.75,
                 chord: Union[list, np.ndarray, float]=0.117,
                 angpitch: Union[list, np.ndarray, float]=0,
                 airfoil: Union[list, np.ndarray, str]='NACA0015',
                 tip_percent: float=0.5,
                 Z: int=7) -> None:
        '''Initializes the turbine with the provided input values.
        
        **parameters**:
    
            :param ttype: string specifying turbine type; either 'Wells' or 'Impulse'. Default is
            'Wells'.
            :param N: number of blade elements (discretization). Default is 50.
            :param rcas: casing radius. Default is 0.25.
            :param hub_to_tip_ratio: hub to tip ratio value. Default is 0.75.
            :param chord: chord value. It is possible to provide an array of length N
            establishing the chord value at each blade element. Default is 0.117, which means
            that the chord is constant along the radius.
            :param angpitch: angular pitch value. It is possible to provide an array of length N
            establishing the angular pitch value at each blade element. Default is 0, which means
            that no pitch is given to the elements.
            :param airfoil: airfoil geometry that corresponds to chord. It is possible to provide
            an array of length N establishing the airfoil geometry at each blade element. Default is
            NACA0015, which means that the whole blade is considered to own a single geometry,
            namely the one of the NACA0015.
            :param tip_percent: tip gap value, in terms of chord percentage. Default is 0.5%.
            :param Z: number of blades. Default is 7.        
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------DECLARING ATTRIBUTES----------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------# 
        
        # Conditional for determining which turbine type is instantiated.
        # If a 'Wells' turbine is instantiated:
        if ttype == "Wells":
            #####
            ##### Setting provided values to attributes.
            #####            
            # Type of turbine.
            self._ttype = 'Wells'
            # Casing radius of turbine.
            self._rcas = rcas
            # Hub-to-tip ratio.
            self._hub_to_tip_ratio = hub_to_tip_ratio
            # Tip percentage.
            self._tip_percent = tip_percent
            # Blade number.
            self._Z = Z
            #####
            ##### Setting calculated values to attributes.
            #####                                 
            # Setting tip radius.
            self._rtip = self.rcas #- self.tip_c
            # Setting hub radius.
            self._rhub = self.rtip*hub_to_tip_ratio
            # Setting differential radial element.
            self._dr = (self._rtip - self._rhub)/N            
            # Setting radial array.
            self._r = np.array([self.rhub + i*self._dr for i in range(N + 1)])  
            # Chord.
            # The conditional structure is intended to cast a single float input value to an array
            # of length equal to the number of blade elements with entries coinciding with the single
            # float value, or to set the 'chord' property of the object equal to the input array or list
            # instead.
            if type(chord) == float:
                self._chord = np.array([chord for _ in self.r])
            elif type(chord) == list:
                self._chord = np.array(chord)
            else:
                self._chord = chord
            # Pitch.
            # The conditional structure is intended to cast a single float input value to an array
            # of length equal to the number of blade elements with entries coinciding with the single
            # float value, or to set the 'angpitch' property of the object equal to the input array or list
            # instead.
            if type(angpitch) == float:
                self._angpitch = np.array([angpitch for _ in self.r])
            elif type(angpitch) == list:
                self._angpitch = np.array(angpitch)
            else:
                self._angpitch = angpitch                
            # Airfoil.
            # The conditional structure is intended to cast a single string input value to a
            # list of length equal to the number of blade elements coinciding with the single string
            # value, or to set the 'airfoil' property of the object equal to the input array or list
            # instead.
            if type(airfoil) == str:
                self._airfoil = np.array([airfoil for _ in self.r])
            elif type(airfoil) == list:
                self._airfoil = airfoil
            else:
                self._airfoil = list(airfoil)
            # Airfoil coords and differential, element-wise areas.
            if type(airfoil) == str:
                # Instantiating airfoil coordinates file name.
                filestr = airfoil + ".txt"
                # Checking for existence of airfoil file.
                if filestr in os.listdir("COORDS/"):
                    # If the file exists, then load coordinates.
                    self._airfoil_coords = [np.loadtxt(open("/".join(["COORDS", filestr])).readlines(), usecols=(0, 1), delimiter="\t", unpack=True) for _ in self.r]
                else:
                    # Otherwise, set coordinates to [[0], [0]] point.
                    self._airfoil_coords = [[[0], [0]] for _ in self.r]                    
            else:
                # Instantiating list for storing airfoils' coordinates.
                self._airfoil_coords = list()
                for _ in airfoil:
                    # Instantiating airfoil coordinates file name.
                    filestr = _ + ".txt"
                    # Checking for existence of airfoil file.
                    if filestr in os.listdir("COORDS/"):   
                        # If the file exists, then load coordinates.
                        self._airfoil_coords.append(np.loadtxt(open("/".join(["COORDS", filestr])).readlines(), usecols=(0, 1), delimiter="\t", unpack=True))
                    else:
                        # Otherwise, set coordinates to [[0], [0]] point.
                        self._airfoil_coords.append([[0], [0]])
            # Rescaling airfoils according to elements' chord values.
            if type(self._chord) == np.ndarray:
                for e, _ in enumerate(self._airfoil_coords):
                    _[0] *= self._chord[e]
                    _[1] *= self._chord[e]
            else:
                for e, _ in enumerate(self._airfoil_coords):
                    _[0] *= self._chord
                    _[1] *= self._chord
            # Computing differential, element-wise areas by calculating the perimeter of the airfoils from the arclength and multiplying them with dr.
            self._dA = np.array([sum(np.sqrt(_[0]**2 + _[1]**2))*self._dr for _ in self._airfoil_coords])
            # Setting tip percent value.
            self._tip_c = self.tip_percent*self.chord[-1]*0.01
            # Setting pitch following 1995Raghunatan.
            self._pitch = 2*np.pi*self.rtip/self.Z - self.chord[-1]
            # Setting aspect ratio.
            self._AR = (self.rtip - self.rhub)/self.chord[-1]              
            # Setting solidity array.
            self._sigma = self._Z*self._chord/(2*np.pi*self._r)
        # If an 'Impulse' turbine is instantiated.
        elif ttype == "Impulse":
            # TODO: Not implemented yet. Raising 'NotImplementedError'.
            raise NotImplementedError()

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------DECLARING PROPERTIES-----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    # Declaration of get property rcas for casing radius.
    @property
    def rcas(self) -> float:
        return self._rcas
    # Declaration of get property hub_to_tip_ratio for hub-to-tip ratio.
    @property
    def hub_to_tip_ratio(self) -> float:
        return self._hub_to_tip_ratio
    # Declaration of get property chord for chord.
    @property
    def chord(self) -> Union[np.ndarray, float]:
        return self._chord
    # Declaration of get property chord for angular pitch.
    @property
    def angpitch(self) -> Union[np.ndarray, float]:
        return self._angpitch    
    # Declaration of get property airfoil for airfoil geometry.    
    @property
    def airfoil(self) -> Union[np.ndarray, str]:
        return self._airfoil    
    # Declaration of get property airfoil_coords for airfoil coordinates.
    def airfoil_coords(self) -> np.ndarray:
        return self._airfoil_coords
    # Declaration of get property dA for differential, element-wise areas.
    @property
    def dA(self) -> np.ndarray:
        return self._dA
    # Declaration of get property tip_percent for tip percent.
    @property
    def tip_percent(self) -> float:
        return self._tip_percent
    # Declaration of get property Z for number of blades.
    @property
    def Z(self) -> int:
        return self._Z
    # Declaration of get property tip_c for tip clearance.
    @property
    def tip_c(self) -> float:
        return self._tip_c
    # Declaration of get property rtip for tip radius.
    @property
    def rtip(self) -> float:
        return self._rtip
    # Declaration of get property rhub for hub radius.
    @property
    def rhub(self) -> float:
        return self._rhub
    # Declaration of get property pitch for pitch.
    @property
    def pitch(self) -> float:
        return self._pitch
    # Declaration of get property AR for aspect ratio.
    @property
    def AR(self) -> float:
        return self._AR
    # Declaration of get property r for radial array.
    @property
    def r(self) -> np.ndarray:
        return self._r
    # Declaration of get property sigma for solidity.
    @property
    def sigma(self) -> float:
        return self._sigma
    # Declaration of get property dr for radial inter-element spacing.
    @property
    def dr(self) -> float:
        return self._dr
       
    # Declaration of set property rcas for casing radius.
    @rcas.setter 
    def rcas(self, value: float) -> None:
        self.__init__(ttype=self._ttype, rcas=value, hub_to_tip_ratio=self._hub_to_tip_ratio,
                     chord=self._chord, angpitch=self._angpitch, airfoil=self._airfoil, tip_percent=self._tip_percent, Z=self._Z)
        return
    # Declaration of set property hub_to_tip_ratio for hub-to-tip ratio.
    @hub_to_tip_ratio.setter
    def hub_to_tip_ratio(self, value :float) -> None:
        self.__init__(ttype=self._ttype, rcas=self._rcas, hub_to_tip_ratio=value,
                     chord=self._chord, angpitch=self._angpitch, airfoil=self._airfoil, tip_percent=self._tip_percent, Z=self._Z)
        return
    # Declaration of set property chord for chord.
    @chord.setter
    def chord(self, value: Union[list, np.ndarray, float]) -> None:
        self.__init__(ttype=self._ttype, rcas=self._rcas, hub_to_tip_ratio=self._hub_to_tip_ratio,
                     chord=value, angpitch=self._angpitch, airfoil=self._airfoil, tip_percent=self._tip_percent, Z=self._Z)
        return
    # Declaration of set property chord for angular pitch.
    @angpitch.setter
    def angpitch(self, value: Union[list, np.ndarray, float]) -> None:
        self.__init__(ttype=self._ttype, rcas=self._rcas, hub_to_tip_ratio=self._hub_to_tip_ratio,
                     chord=self._chord, angpitch=value, airfoil=self._airfoil, tip_percent=self._tip_percent, Z=self._Z)
        return    
    # Declaration of set property airfoil for airfoil geometry.
    @airfoil.setter
    def airfoil(self, value: Union[list, np.ndarray, str]) -> None:
        self.__init__(ttype=self._ttype, rcas=self._rcas, hub_to_tip_ratio=self._hub_to_tip_ratio,
                     chord=self._chord, angpitch=self._angpitch, airfoil=value, tip_percent=self._tip_percent, Z=self._Z)
        return
    # Declaration of set property tip_percent for tip percent.
    @tip_percent.setter
    def tip_percent(self, value: float) -> None:
        self.__init__(ttype=self._ttype, rcas=self._rcas, hub_to_tip_ratio=self._hub_to_tip_ratio,
                     chord=self._chord, angpitch=self._angpitch, airfoil=self._airfoil, tip_percent=value, Z=self._Z)
        return
    # Declaration of set property Z for number of blades.
    @Z.setter
    def Z(self, value: int) -> None:
        self.__init__(ttype=self._ttype, rcas=self._rcas, hub_to_tip_ratio=self._hub_to_tip_ratio,
                     chord=self._chord, angpitch=self._angpitch, airfoil=self._airfoil, tip_percent=self._tip_percent, Z=value)
        return
    
#--------------------------------------------------------------------------------------------------------------------#
#---------------------------------------------------CLASS flow_params------------------------------------------------#             
#--------------------------------------------------------------------------------------------------------------------#     
class flow_params:
    '''Flow parameters entering a turbine stage.'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------__init__() METHOD------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def __init__(self,
                 p: float=101325,
                 T: float=288.15,
                 R: float=287.058,
                 nu: float=1.81e-5) -> None:
        '''Initializes a given instance of the input parameters entering a turbine.
        
        **parameters**, **return**, **return types**:
    
            :param p: atmospheric pressure in [Pa]. Default value is 101325 [Pa].
            :param T: ambient temperature in [K]. Default value is 288.15 [K].
            :param R: gas constant in [J/(kg·K)]. Default value is 287.058 [J/(kg·K)].
            :param nu: gas viscosity in [kg/(m·s)]. Default value is 1.81e-5 [kg/(m·s)].       
        '''        
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------DECLARING ATTRIBUTES-----------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
    
        # Atmospheric pressure.
        self._p = p
        # Ambient temperature.
        self._T = T
        # Gas constant.
        self._R = R
        # Fluid viscosity.
        self._nu = nu
        # Calculating density assuming ideal gas. Setting it to 'rho' attribute.
        self._rho = p/(T*R)

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------DECLARING PROPERTIES-----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#                
        
    # Declaration of get property p for atmospheric pressure.
    @property
    def p(self) -> float:
        return self._p
    # Declaration of get property T for ambient temperature.
    @property
    def T(self) -> float:
        return self._T
    # Declaration of get property R for gas (air) constant.
    @property
    def R(self) -> float:
        return self._R
    # Declaration of get property nu for gas (air) viscosity.
    @property
    def nu(self) -> float:
        return self._nu
    # Declaration of get property rho for gas (air) density.
    @property
    def rho(self) -> float:
        return self._rho

    # Declaration of set property p for atmospheric pressure.
    @p.setter 
    def p(self, value: float) -> None:
        self.__init__(p=value, T=self._T, R=self._R, nu=self._nu)
        return
    # Declaration of set property T for ambient temperature.
    @T.setter
    def T(self, value: float) -> None:
        self.__init__(p=self._P, T=value, R=self._R, nu=self._nu)
        return
    # Declaration of set property R for gas (air) constant.
    @R.setter
    def R(self, value: float) -> None:
        self.__init__(p=self._P, T=self._T, R=value, nu=self._nu)
        return
    # Declaration of set property nu for gas (air) viscosity.
    @nu.setter
    def nu(self, value: float) -> None:
        self.__init__(p=self._P, T=self._T, R=self._R, nu=value)
        return
    
#--------------------------------------------------------------------------------------------------------------------#
#------------------------------------------------CLASS velocity_triangle---------------------------------------------#             
#--------------------------------------------------------------------------------------------------------------------#
class velocity_triangle:
    '''Spanwise velocity triangle of a turbine stage.'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------__init__() METHOD------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def __init__(self, N: int=50) -> None:
        '''Initializes a given instance of a velocity triangle.
        
        **parameters**, **return**, **return type**:
            :param N: number of blade elements (discretization). Default is 50.      
        '''        
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------DECLARING ATTRIBUTES-----------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Absolute angle.
        self._alpha = np.zeros(N + 1)
        # Axial component of absolute velocity.
        self._vx = np.zeros(N + 1)
        # Tangential component of absolute velocity.
        self._vtheta = np.zeros(N + 1)
        # Absolute velocity magnitude.
        self._v = np.zeros(N + 1)
        # Tangential velocity.
        self._U = np.zeros(N + 1)
        # Axial component of relative velocity.
        self._wx = np.zeros(N + 1)
        # Tangential component of relative velocity.
        self._wtheta = np.zeros(N + 1)
        # Relative velocity magnitude.
        self._w = np.zeros(N + 1)
        # Relative angle.
        self._beta = np.zeros(N + 1)
        # Flow coefficient.
        self._phi = np.zeros(N + 1)
        # Differential flow-rate.
        self._dq = np.zeros(N + 1)
        # Flow-rate.
        self._q = 0
        # Non-dimensional flow-rate.
        self._Phi = 0

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------DECLARING PROPERTIES-----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#                
        
    # Declaration of get property alpha for absolute velocity angle.
    @property
    def alpha(self) -> np.ndarray:
        return self._alpha
    # Declaration of get property vx for axial component of absolute velocity.
    @property
    def vx(self) -> np.ndarray:
        return self._vx
    # Declaration of get property vtheta for tangential component of absolute velocity.
    @property
    def vtheta(self) -> np.ndarray:
        return self._vtheta
    # Declaration of get property v for magnitude of absolute velocity.
    @property
    def v(self) -> np.ndarray:
        return self._v
    # Declaration of get property U for tangential velocity.
    @property
    def U(self) -> np.ndarray:
        return self._U
    # Declaration of get property wx for axial component of relative velocity.
    @property
    def wx(self):
        return self._wx
    # Declaration of get property wtheta for tangential component of relative velocity.
    @property
    def wtheta(self) -> np.ndarray:
        return self._wtheta
    # Declaration of get property w for magnitude of relative velocity.
    @property
    def w(self) -> np.ndarray:
        return self._w
    # Declaration of get property beta for relative velocity angle.
    @property
    def beta(self) -> np.ndarray:
        return self._beta
    # Declaration of get property phi for flow coefficient.
    @property
    def phi(self) -> np.ndarray:
        return self._phi
    # Declaration of get property dq for differential flow-rate value.
    @property
    def dq(self) -> np.ndarray:
        return self._dq
    # Declaration of get property q for flow-rate value.
    @property
    def q(self) -> float:
        return self._q
    # Declaration of get property Phi for dimensionless flow-rate value.
    @property
    def Phi(self) -> float:
        return self._Phi
    
    # Declaration of set property alpha for absolute velocity angle.
    @alpha.setter 
    def alpha(self, value: np.ndarray) -> None:
        assert len(value) == len(self._alpha), 'Provide an array of size ' + len(self._alpha) + '; current array has size ' + str(len(value))
        self._alpha = value
        return
    # Declaration of set property vx for axial component of absolute velocity.
    @vx.setter
    def vx(self, value: np.ndarray) -> None:
        assert len(value) == len(self._vx), 'Provide an array of size ' + len(self._vx) + '; current array has size ' + str(len(value))
        self._vx = value
        return
    # Declaration of set property vtheta for tangential component of absolute velocity.
    @vtheta.setter
    def vtheta(self, value: np.ndarray) -> None:
        assert len(value) == len(self._vtheta), 'Provide an array of size ' + len(self._vtheta) + '; current array has size ' + str(len(value))
        self._vtheta = value
        return
    # Declaration of set property v for magnitude of absolute velocity.
    @v.setter
    def v(self, value: np.ndarray) -> None:
        assert len(value) == len(self._v), 'Provide an array of size ' + len(self._v) + '; current array has size ' + str(len(value))
        self._v = value
        return
    # Declaration of set property U for tangential velocity.
    @U.setter 
    def U(self, value: np.ndarray) -> None:
        assert len(value) == len(self._U), 'Provide an array of size ' + len(self._U) + '; current array has size ' + str(len(value))
        self._U = value
        return
    # Declaration of set property wx for axial component of relative velocity.
    @wx.setter
    def wx(self, value: np.ndarray) -> None:
        assert len(value) == len(self._wx), 'Provide an array of size ' + len(self._wx) + '; current array has size ' + str(len(value))
        self._wx = value
        return
    # Declaration of set property wtheta for tangential component of relative velocity.
    @wtheta.setter
    def wtheta(self, value: np.ndarray) -> None:
        assert len(value) == len(self._wtheta), 'Provide an array of size ' + len(self._wtheta) + '; current array has size ' + str(len(value))
        self._wtheta = value
        return
    # Declaration of set property w for magnitude of relative velocity.
    @w.setter
    def w(self, value: np.ndarray) -> None:
        assert len(value) == len(self._w), 'Provide an array of size ' + len(self._w) + '; current array has size ' + str(len(value))
        self._w = value
        return        
    # Declaration of set property beta for relative velocity angle.
    @beta.setter
    def beta(self, value: np.ndarray) -> None:
        assert len(value) == len(self._beta), 'Provide an array of size ' + len(self._beta) + '; current array has size ' + str(len(value))
        self._beta = value
        return
    # Declaration of set property phi for flow coefficient.
    @phi.setter
    def phi(self, value: np.ndarray) -> None:
        assert len(value) == len(self._phi), 'Provide an array of size ' + len(self._phi) + '; current array has size ' + str(len(value))
        self._phi = value
        return
    # Declaration of set property dq for differential flow-rate value.
    @dq.setter
    def dq(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dq), 'Provide an array of size ' + len(self._dq) + '; current array has size ' + str(len(value))
        self._dq = value
        return
    # Declaration of set property q for flow-rate value.
    @q.setter
    def q(self, value: float) -> None:        
        self._q = value
        return
    # Declaration of set property Phi for dimensionless flow-rate value.
    @Phi.setter
    def Phi(self, value: float) -> None:
        self._Phi = value
        return
    
#--------------------------------------------------------------------------------------------------------------------#
#---------------------------------------------------CLASS load_coefs-------------------------------------------------#             
#--------------------------------------------------------------------------------------------------------------------#
class load_coefs:
    '''Spanwise load coefficients on a turbine stage.'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------__init__() METHOD------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def __init__(self, N: int=50):
        '''Initializes a given instance of a class containing loads and coefficients.
        
        **parameters**, **return**, **return type**:
            :param N: number of blade elements (discretization). Default is 50.
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------DECLARING ATTRIBUTES-----------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Reynolds number.
        self._Re = np.zeros(N + 1)
        # Lift coefficient.
        self._cl = np.zeros(N + 1)
        # Differential lift force.
        self._dl = np.zeros(N + 1)
        # Lift-state coefficient.
        self._clstate = ["" for _ in self._cl]
        # Drag coefficient.
        self._cd = np.zeros(N + 1)
        # Differential drag force.
        self._dd = np.zeros(N + 1)
        # Induced drag coefficient.
        self._cdtc = np.zeros(N + 1)
        # Differential induced drag force.
        self._ddtc = np.zeros(N + 1)        
        # Axial force coefficient.
        self._cx = np.zeros(N + 1)
        # Differential axial force.
        self._dx = np.zeros(N + 1)        
        # Tangential force coefficient.
        self._ctheta = np.zeros(N + 1)
        # Differential tangential force.
        self._dtheta = np.zeros(N + 1)        
        # Interference-corrected axial force coefficient.
        self._cxCIF = np.zeros(N + 1)
        # Differential interference-corrected axial force.
        self._dxCIF = np.zeros(N + 1)        
        # Interference-corrected tangential force coefficient.
        self._cthetaCIF = np.zeros(N + 1)
        # Differential interference-corrected tangential force.
        self._dthetaCIF = np.zeros(N + 1)        
        # Input coefficient.
        self._cinput = 0
        # Torque coefficient.
        self._ctau = 0

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------DECLARING PROPERTIES-----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#                
        
    # Declaration of get property Re for Reynolds numbers.
    @property
    def Re(self) -> np.ndarray:
        return self._Re
    # Declaration of get property cl for lift coefficients.
    @property
    def cl(self) -> np.ndarray:
        return self._cl
    # Declaration of get property dl for differential lift forces.
    @property
    def dl(self) -> np.ndarray:
        return self._dl    
    # Declaration of get property clstate for lift coefficient states.
    @property
    def clstate(self) -> np.ndarray:
        return self._clstate   
    # Declaration of get property cd for drag coefficients.
    @property
    def cd(self) -> np.ndarray:
        return self._cd
    # Declaration of get property dd for differential drag forces.
    @property
    def dd(self) -> np.ndarray:
        return self._dd    
    # Declaration of get property cdtc for induced drag coefficients.
    @property
    def cdtc(self) -> np.ndarray:
        return self._cdtc
    # Declaration of get property ddtc for differential induced drag forces.
    @property
    def ddtc(self) -> np.ndarray:
        return self._ddtc    
    # Declaration of get property cx for axial coefficients.
    @property
    def cx(self) -> np.ndarray:
        return self._cx
    # Declaration of get property dx for differential axial forces.
    @property
    def dx(self) -> np.ndarray:
        return self._dx    
    # Declaration of get property ctheta for tangential coefficients.
    @property
    def ctheta(self) -> np.ndarray:
        return self._ctheta
    # Declaration of get property dtheta for differential tangential forces.
    @property
    def dtheta(self) -> np.ndarray:
        return self._dtheta    
    # Declaration of get property cxCIF for corrected axial coefficients.
    @property
    def cxCIF(self) -> np.ndarray:
        return self._cxCIF
    # Declaration of get property dxCIF for corrected differential axial forces.
    @property
    def dxCIF(self) -> np.ndarray:
        return self._dxCIF
    # Declaration of get property cthetaCIF for corrected tangential coefficients.    
    @property
    def cthetaCIF(self) -> np.ndarray:
        return self._cthetaCIF
    # Declaration of get property dthetaCIF for corrected differential tangential coefficients.    
    @property
    def dthetaCIF(self) -> np.ndarray:
        return self._dthetaCIF    
    # Declaration of get property cinput for input coefficients.
    @property
    def cinput(self) -> np.ndarray:
        return self._cinput
    # Declaration of get property ctau for torque coefficients.
    @property
    def ctau(self) -> np.ndarray:
        return self._ctau
    
    # Declaration of set property Re for Reynolds number.
    @Re.setter 
    def Re(self, value: np.ndarray) -> None:
        assert len(value) == len(self._Re), 'Provide an array of size ' + len(self._Re) + '; current array has size ' + str(len(value))
        self._Re = value
        return
    # Declaration of set property cl for lift coefficients.
    @cl.setter
    def cl(self, value: np.ndarray) -> None:
        assert len(value) == len(self._cl), 'Provide an array of size ' + len(self._cl) + '; current array has size ' + str(len(value))
        self._cl = value
        return
    # Declaration of set property dl for lift forces.
    @dl.setter
    def dl(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dl), 'Provide an array of size ' + len(self._dl) + '; current array has size ' + str(len(value))
        self._dl = value
        return    
    # Declaration of set property clstate for lift coefficient states.
    @clstate.setter
    def clstate(self, value: np.ndarray) -> None:
        assert len(value) == len(self._clstate), 'Provide an array of size ' + len(self._cl) + '; current array has size ' + str(len(value))
        self._clstate = value
        return    
    # Declaration of set property cd for drag coefficients.
    @cd.setter
    def cd(self, value: np.ndarray) -> None:
        assert len(value) == len(self._cd), 'Provide an array of size ' + len(self._cd) + '; current array has size ' + str(len(value))
        self._cd = value
        return
    # Declaration of set property dd for drag forces.
    @dd.setter
    def dd(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dd), 'Provide an array of size ' + len(self._dd) + '; current array has size ' + str(len(value))
        self._dd = value
        return    
    # Declaration of set property cdtc for induced drag coefficients.
    @cdtc.setter
    def cdtc(self, value: np.ndarray) -> None:
        assert len(value) == len(self._cdtc), 'Provide an array of size ' + len(self._cdtc) + '; current array has size ' + str(len(value))
        self._cdtc = value
        return
    # Declaration of set property ddtc for induced drag loads.
    @ddtc.setter
    def ddtc(self, value: np.ndarray) -> None:
        assert len(value) == len(self._ddtc), 'Provide an array of size ' + len(self._ddtc) + '; current array has size ' + str(len(value))
        self._ddtc = value
        return    
    # Declaration of set property cx for axial coefficients.
    @cx.setter
    def cx(self, value: np.ndarray) -> None:
        assert len(value) == len(self._cx), 'Provide an array of size ' + len(self._cx) + '; current array has size ' + str(len(value))
        self._cx = value
        return
    # Declaration of set property dx for axial loads.
    @dx.setter
    def dx(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dx), 'Provide an array of size ' + len(self._dx) + '; current array has size ' + str(len(value))
        self._dx = value
        return    
    # Declaration of set property ctheta for tangential coefficients.
    @ctheta.setter 
    def ctheta(self, value: np.ndarray) -> None:
        assert len(value) == len(self._ctheta), 'Provide an array of size ' + len(self._ctheta) + '; current array has size ' + str(len(value))
        self._ctheta = value
        return
    # Declaration of set property dtheta for tangential coefficients.
    @dtheta.setter 
    def dtheta(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dtheta), 'Provide an array of size ' + len(self._dtheta) + '; current array has size ' + str(len(value))
        self._dtheta = value
        return    
    # Declaration of set property cxCIF for corrected axial coefficients.
    @cxCIF.setter
    def cxCIF(self, value: np.ndarray) -> None:
        assert len(value) == len(self._cxCIF), 'Provide an array of size ' + len(self._cxCIF) + '; current array has size ' + str(len(value))
        self._cxCIF = value
        return
    # Declaration of set property dxCIF for corrected axial coefficients.
    @dxCIF.setter
    def dxCIF(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dxCIF), 'Provide an array of size ' + len(self._dxCIF) + '; current array has size ' + str(len(value))
        self._dxCIF = value
        return    
    # Declaration of set property cthetaCIF for corrected tangential coefficients.
    @cthetaCIF.setter
    def cthetaCIF(self, value: np.ndarray) -> None:
        assert len(value) == len(self._cthetaCIF), 'Provide an array of size ' + len(self._cthetaCIF) + '; current array has size ' + str(len(value))
        self._cthetaCIF = value
        return
    # Declaration of set property dthetaCIF for corrected tangential coefficients.
    @dthetaCIF.setter
    def dthetaCIF(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dthetaCIF), 'Provide an array of size ' + len(self._dthetaCIF) + '; current array has size ' + str(len(value))
        self._dthetaCIF = value
        return    
    # Declaration of set property cinput for input coefficients.
    @cinput.setter
    def cinput(self, value: np.ndarray) -> None:
        self._cinput = value
        return  
    # Declaration of set property ctau for torque coefficients.
    @ctau.setter
    def ctau(self, value: np.ndarray) -> None:
        self._ctau = value
        return
    
#--------------------------------------------------------------------------------------------------------------------#
#---------------------------------------------------CLASS power_output-----------------------------------------------#             
#--------------------------------------------------------------------------------------------------------------------#
class power_output:
    '''Energetic outpus of a turbine stage.'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------__init__() METHOD------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def __init__(self, N: int=50):
        '''Initializes a given instance of a class containing the energetic outputs of a turbine stage.
        
        **parameters**, **return**, **return type**:
            :param N: number of blade elements (discretization). Default is 50.
            :return: instance of 'power_output' class.
            :rtype: type(power_output), obj.        
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------DECLARING ATTRIBUTES-----------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Radially-varying differential torque.
        self._difftau = np.zeros(N + 1)
        # Radially-varying differential power.
        self._diffpower = np.zeros(N + 1)
        # Radially-varying static-to-static pressure drop.
        self._dp_st_to_st = np.zeros(N + 1)
        # Radially-varying total-to-static pressure drop.
        self._dp_tot_to_st = np.zeros(N + 1)
        # Radially-varying total-to-total pressure drop.
        self._dp_tot_to_tot = np.zeros(N + 1)
        # Radially-varying viscous losses.
        self._dpvisc = np.zeros(N + 1)
        # Radially-varying kinetic losses.
        self._dpk = np.zeros(N + 1)
        # Radially-varying efficiency.
        self._deff = np.zeros(N + 1)
        # Integrated torque value.
        self._tau = 0
        # Integrated power.
        self._power = 0
        # Non-dimensional power.
        self._Pi = 0
        # Integrated static-to-static pressure drop.
        self._p_st_to_st = 0
        # Integrated total-to-static pressure drop.
        self._p_tot_to_st = 0
        # Non-dimensional integrated total-to-static pressure drop.
        self._Psi = 0
        # Integrated total-to-total pressure drop.
        self._p_tot_to_tot = 0
        # Integrated viscous losses.
        self._pvisc = 0
        # Integrated kinetic losses.
        self._pk = 0
        # Integrated efficiency.
        self._eff = 0
        # Efficiency coming from non-dimensional quantities.
        self._Eta = 0

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------DECLARING PROPERTIES-----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#                
        
    # Declaration of get property difftau for differential torque.
    @property
    def difftau(self) -> np.ndarray:
        return self._difftau
    # Declaration of get property diffpower for differential power.
    @property
    def diffpower(self) -> np.ndarray:
        return self._diffpower
    # Declaration of get property dp_st_to_st for differential static-to-static pressure-drop.
    @property
    def dp_st_to_st(self) -> np.ndarray:
        return self._dp_st_to_st
    # Declaration of get property dp_tot_to_st for differential total-to-static pressure-drop.
    @property
    def dp_tot_to_st(self) -> np.ndarray:
        return self._dp_tot_to_st
    # Declaration of get property dp_tot_to_tot for differential total-to-total pressure-drop.
    @property
    def dp_tot_to_tot(self) -> np.ndarray:
        return self._dp_tot_to_tot
    # Declaration of get property dpvisc for differential pressure-drop due to viscous losses.
    @property
    def dpvisc(self) -> np.ndarray:
        return self._dpvisc
    # Declaration of get property dpk for differential pressure-drop due to kinetic losses.
    @property
    def dpk(self) -> np.ndarray:
        return self._dpk
    # Declaration of get property deff for differential efficiency.
    @property
    def deff(self) -> np.ndarray:
        return self._deff
    # Declaration of get property tau for torque.
    @property
    def tau(self) -> float:
        return self._tau
    # Declaration of get property power for power.
    @property
    def power(self) -> float:
        return self._power
    # Declaration of get property Pi for dimensionless power.
    @property
    def Pi(self) -> float:
        return self._Pi
    # Declaration of get property p_st_to_st for static-to-static pressure-drop.
    @property
    def p_st_to_st(self) -> float:
        return self._p_st_to_st
    # Declaration of get property p_tot_to_st for total-to-static pressure-drop.
    @property
    def p_tot_to_st(self) -> float:
        return self._p_tot_to_st
    # Declaration of get property p_tot_to_tot for total-to-total pressure-drop.
    @property
    def p_tot_to_tot(self) -> float:
        return self._p_tot_to_tot
    # Declaration of get property Psi for dimensionless total-to-total pressure-drop.
    @property
    def Psi(self) -> float:
        return self._Psi
    # Declaration of get property pvisc for pressure-drop due to viscous losses.
    @property
    def pvisc(self) -> float:
        return self._pvisc
    # Declaration of get property pk for pressure-drop due to kinetic losses.
    @property
    def pk(self) -> float:
        return self._pk
    # Declaration of get property eff for efficiency.
    @property
    def eff(self) -> float:
        return self._eff
    # Declaration of get property Eta for efficiency coming from dimensionless variables.
    @property
    def Eta(self) -> float:
        return self._Eta
    
    # Declaration of set property difftau for differential torque.    
    @difftau.setter 
    def difftau(self, value: np.ndarray) -> None:
        assert len(value) == len(self._difftau), 'Provide an array of size ' + len(self._difftau) + '; current array has size ' + str(len(value))
        self._difftau = value
        return
    # Declaration of set property diffpower for differential power.
    @diffpower.setter
    def diffpower(self, value: np.ndarray) -> None:
        assert len(value) == len(self._diffpower), 'Provide an array of size ' + len(self._diffpower) + '; current array has size ' + str(len(value))
        self._diffpower = value
        return
    # Declaration of set property dp_st_to_st for differential static-to-static pressure-drop.
    @dp_st_to_st.setter
    def dp_st_to_st(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dp_st_to_st), 'Provide an array of size ' + len(self._dp_st_to_st) + '; current array has size ' + str(len(value))
        self._dp_st_to_st = value
        return
    # Declaration of set property dp_tot_to_st for differential total-to-static pressure-drop.
    @dp_tot_to_st.setter
    def dp_tot_to_st(self, value: np.ndarray) -> None:
        assert len(value) ==len(self._dp_tot_to_st), 'Provide an array of size ' + len(self._dp_tot_to_st) + '; current array has size ' + str(len(value))
        self._dp_tot_to_st = value
        return
    # Declaration of set property dp_tot_to_tot for differential total-to-total pressure-drop.
    @dp_tot_to_tot.setter 
    def dp_tot_to_tot(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dp_tot_to_tot), 'Provide an array of size ' + len(self._dp_tot_to_tot) + '; current array has size ' + str(len(value))
        self._dp_tot_to_tot = value
        return
    # Declaration of set property dpvisc for differential pressure-drop due to viscous losses.
    @dpvisc.setter
    def dpvisc(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dpvisc), 'Provide an array of size ' + len(self._dpvisc) + '; current array has size ' + str(len(value))
        self._dpvisc = value
        return
    # Declaration of set property dpk for differential pressure-drop due to kinetic losses.    
    @dpk.setter
    def dpk(self, value: np.ndarray) -> None:
        assert len(value) == len(self._dpk), 'Provide an array of size ' + len(self._dpk) + '; current array has size ' + str(len(value))
        self._dpk = value
        return   
    # Declaration of set property deff for differential efficiency.    
    @deff.setter
    def deff(self, value: np.ndarray) -> None:
        assert len(value) == len(self._deff), 'Provide an array of size ' + len(self._deff) + '; current array has size ' + str(len(value))
        self._deff = value
        return
    # Declaration of set property tau for torque.
    @tau.setter 
    def tau(self, value: float) -> None:        
        self._tau = value
        return
    # Declaration of set property power for power.    
    @power.setter
    def power(self, value: float) -> None:        
        self._power = value
        return
    # Declaration of set property Pi for dimensionless power.
    @Pi.setter
    def Pi(self, value: float) -> None:
        self._Pi = value
        return
    # Declaration of set property p_st_to_st for static-to-static pressure-drop.
    @p_st_to_st.setter
    def p_st_to_st(self, value: float) -> None:
        self._p_st_to_st = value
        return
    # Declaration of set property p_tot_to_st for total-to-static pressure-drop.
    @p_tot_to_st.setter
    def p_tot_to_st(self, value: float) -> None:
        self._p_tot_to_st = value
        return
    # Declaration of set property p_tot_to_tot for total-to-total pressure-drop.
    @p_tot_to_tot.setter 
    def p_tot_to_tot(self, value: float) -> None:
        self._p_tot_to_tot = value
        return
    # Declaration of set property Psi for dimensionless total-to-total pressure-drop.    
    @Psi.setter
    def Psi(self, value: float) -> None:
        self._Psi = value
        return
    # Declaration of set property pvisc for pressure-drop due to viscous losses.    
    @pvisc.setter
    def pvisc(self, value: float) -> None:
        self._pvisc = value
        return
    # Declaration of set property pk for pressure-drop due to kinetic losses.    
    @pk.setter
    def pk(self, value: float) -> None:
        self._pk = value
        return       
    # Declaration of set property eff for efficiency.        
    @eff.setter
    def eff(self, value: float) -> None:
        self._eff = value
        return
    # Declaration of set property Eta for efficiency coming from dimensionless variables.
    @Eta.setter
    def Eta(self, value: float) -> None:
        self._Eta = value
        return

#--------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------CLASS turbine_stage-----------------------------------------------#             
#--------------------------------------------------------------------------------------------------------------------#
class turbine_stage:
    '''Sets a turbine stage.'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------__init__() METHOD------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def __init__(self,
                 ttype: str='Wells',
                 omega: float=3600,                 
                 N: int=50,
                 rcas: float=0.25,
                 hub_to_tip_ratio: float=0.75,
                 chord: Union[list, np.ndarray, float]=0.117,
                 angpitch: Union[list, np.ndarray, float]=0,
                 airfoil: Union[str, np.ndarray, float]="NACA0015",
                 tip_percent: float=0.5,
                 Z: int=7,
                 p: float=101325,
                 T: float=288.15,
                 R: float=287.058,
                 nu: float=1.81e-5) -> None:
        '''Initializes a given instance of a class containing the a turbine stage.
        
        **parameters**, **return**, **return type**:              
            :param ttype: string specifying turbine type; either 'Wells' or 'Impulse'. Default is
            'Wells'.
            :param omega: rotational speed of turbine stage, in [RPM]. Default is 3600 [RPM].
            :param N: number of blade elements (discretization). Default is 50.
            :param rcas: casing radius. Default is 0.25.
            :param hub_to_tip_ratio: hub to tip ratio value. Default is 0.75.
            :param chord: chord value. Default is 0.117. It can be an element-wise array.
            :param angpitch: angular pitch value. Default is 0. It can be an element-wise array.
            :param airfoil: geometry of the airfoils of the blade. Default is NACA0015, meaning that
            the airfoil geometry is constant throughout the blade. It can be an element-wise array.
            :param tip_percent: tip gap value, in terms of chord percentage. Default is 0.5%.
            :param Z: number of blades. Default is 7.
            :param p: atmospheric pressure in [Pa]. Default value is 101325 [Pa].
            :param T: ambient temperature in [K]. Default value is 288.15 [K].
            :param R: gas constant in [J/(kg·K)]. Default value is 287.058 [J/(kg·K)]
            :param nu: gas viscosity in [kg/(m·s)]. Default value is 1.81e-5 [kg/(m·s)].
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------DECLARING ATTRIBUTES-----------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Type of turbine.
        self._ttype = ttype
        # Rotational speed.
        self._omega = omega*2*np.pi/60
        # Geometrical parameters.
        self._gp = geom_params(ttype=ttype,
                               N=N,
                               rcas=rcas,
                               hub_to_tip_ratio=hub_to_tip_ratio,
                               chord=chord,
                               angpitch=angpitch,
                               airfoil=airfoil,
                               tip_percent=tip_percent,
                               Z=Z
                              )
        # Flow parameters.
        self._fp = flow_params(p=p,
                               T=T,
                               R=R,
                               nu=nu
                              )
        # Inlet velocity triangle.
        self._it = velocity_triangle(N=N)
        # Outlet velocity triangle.
        self._ot = velocity_triangle(N=N)
        # Coefficients.
        self._coefs = load_coefs(N=N)
        # Energetic output.
        self._eout = power_output(N=N)

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------DECLARING PROPERTIES-----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#        
        
    # Declaration of get property ttype for turbine stage type.
    @property
    def ttype(self) -> str:
        return self._ttype
    # Declaration of get property omega for rotational speed.
    @property
    def omega(self) -> float:
        return self._omega
    # Declaration of get property gp for turbine stage geometrical parameters.
    @property
    def gp(self) -> Union[float, np.ndarray]:
        return self._gp
    # Declaration of get property fp for turbine stage flow parameters.
    @property
    def fp(self) -> float:
        return self._fp
    # Declaration of get property it for turbine stage inlet velocity triangle.
    @property
    def it(self) -> velocity_triangle:
        return self._it
    # Declaration of get property ot for turbine stage outlet velocity triangle.
    @property
    def ot(self) -> velocity_triangle:
        return self._ot
    # Declaration of get property coefs for turbine stage load coefficients.
    @property
    def coefs(self) -> load_coefs:
        return self._coefs
    # Declaration of get property eout for turbine energetic output.
    @property
    def eout(self) -> power_output:
        return self._eout
    
    # Declaration of set property omega for rotational speed.
    @omega.setter
    def omega(self, value) -> None:
        self._omega = value
        return
    # Declaration of set property it for input velocity triangle.
    @it.setter
    def it(self, value) -> None:
        self._it = value
        return
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-----------------------------------------------------BEM() METHOD---------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def BEM(self,
            phi_q_v: Union[list, np.ndarray, float],
            polars: dict,
            inputparam: str='phi',
            mode: str='3DNOAD') -> None:
        '''Applies the momentum element method to the current turbine stage.
        
        **parameters**, **return**, **return type**:                    
            :param phi_q_v: input parameter value for which the BEM method is applied. It can
            be either a flow coefficient value (phi), a flow-rate value (q) or an axial velocity
            value (vx).
            :param polars: dictionary of lift and drag coefficients at different Reynolds numbers for
            a set of distinct airfoils from which to interpolate the lift and drag coefficients
            upon each of the blade elements.
            :param inputparam: string determining the type of input parameter. It can be either
            'phi' (stipulating flow coefficient), 'flowrate' (flow-rate) or 'axialvel'
            (axial velocity). Default is 'phi'.
            :param mode: string determining the model of BEM to be applied. It can be either
            '3DNOAD' (semi-3D BEM model without actuator disk) or '3DAD' (semi-3D BEM model with
            actuator disk). Default is '3DNOAD'.     
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #------------------------------------------------------ASSERTIONS----------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#        
        
        # Asserts whether the 'mode' input variable is valid. If not, it raises an error message.
        assert mode in ["3DNOAD", "3DAD"], "Provide a valid mode for applying the BEM method, either '3DNOAD' or '3DAD''; current value is " + str(mode)
        
        #--------------------------------------------------------------------------------------------------------------------#
        #---------------------------------------------------------BODY-------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#        
        
        # Conditional determining which of the BEM models is to be applied.
        if mode == '3DNOAD':
            # Call to external function 'BEM_wells_noad'.
            self.BEM_wells_noad(phi_q_v, self.omega, polars, inputparam, refvel='it')
        elif mode == '3DAD':
            # Call to external function 'BEM_wells_ads'.
            self.BEM_wells_ad(phi_q_v, self.omega, polars, inputparam, refvel='it')
        
        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------------RETURN------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Return statement.
        return
    
    #--------------------------------------------------------------------------------------------------------------------#
    #------------------------------------------------BEM_wells_noad() METHOD---------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def BEM_wells_noad(self,
                       phi_q_v: Union[list, np.ndarray, float],
                       omega: float,
                       polars: dict,
                       inputparam: str="vx",
                       refvel: str='it') -> None:
        '''Applies the blade-element momentum theory (BEM) to a Wells-type turbine stage.

        The BEM theory is applied as formulated in the Ph.D. dissertation thesis of Ciappi (2021),
        available in the link https://flore.unifi.it/handle/2158/1245178#.YWXWWRpBzIV.

        - **parameters**, **returns**, **return types**

        :param ts: instance of the turbine_stage class to which the BEM is to be applied.
        :param phi_q_v: input parameter, either a flow coefficient (phi) a flow-rate value (q) or
        an axial velocity value (vx). It may be either an array-like structure, specifying phi, q
        or v for each blade element, or a single value, in which the provided value is set equal
        to all the blade elements.
        :param omega: rotational speed of the turbine [rad/s].
        :param polars: a dictionary containing the polars of the element airfoils.
        :param inputparam: string representing the type of input provided; it is "phi" for indicating
        a flow coefficient, "q" for a flow-rate, "vx" for an axial velocity or "it" for a pre-established
        inlet velocity triangle. Default is "vx".
        :param refvel: string representing the referential velocity employed for calculating loads
        and velocity-dependent outputs; it is "it" for setting the referential velocity equal to the
        relative velocity magnitude of the inlet velocity triangle, or 'ave' for setting it equal to
        the relative velocity magnitude based on the actuator disk value (the axial component being
        the average of downstream and upstream velocities). Default is "it".
        '''

        #--------------------------------------------------------------------------------------------------------------------#
        #------------------------------------------------------ASSERTIONS----------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Assertion for checking that the input parameter matches an accepted value.
        assert inputparam in ["phi", "q", "vx", "it"], "Provide a valid inputparam mode: 'phi', 'q', 'vx' or 'it'."
        # Assertion for checking that the string representing the referential velocity "refvel" has a valid value.
        assert refvel in ["it", "ave"], "Provide a valid 'refvel' value: 'it' or 'ave'."
        # Conditional for asserting that the input parameter has the same length of the array of blade elements.
        if isinstance(phi_q_v, np.ndarray) or isinstance(phi_q_v, list):
            assert len(phi_q_v) == len(self.gp.r), 'In case the input variable is meant to be an array  provide it so that its length equals the number of blade elements. Input array has a length of ' + str(len(phi_q_v)) + ', and number of blade elements is ' + str(len(self.gp.r)) + '.'
            
        #--------------------------------------------------------------------------------------------------------------------#
        #---------------------------------------------------------BODY-------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#             

        #####
        ##### TYPE CASTING
        #####
        # Changing the type of input data to float, in case it is necessary.
        if isinstance(phi_q_v, np.ndarray):
            # Action in case the input data is in np.ndarray format.
            if "float" not in phi_q_v.dtype.name:
                phi_q_v = phi_q_v.astype(float)
            else:
                pass
        elif isinstance(phi_q_v, list):
            # Action in case the input data is in list format.
            phi_q_v = [float(_) for _ in phi_q_v]
        else:
            # Action in case the input data is in single variable format, but not float (integer possibly).
            if type(phi_q_v) != type(float):
                phi_q_v = float(phi_q_v)

        #####
        ##### UPSTREAM VELOCITY-TRIANGLE CALCULATION
        #####
        # Conditional for checking whether "inputparam" is "it"; in such a case, a pre-established
        # velocity triangle is provided as an input, and it is not necessary to begin the calculations
        # from determining the inlet velocity triangle.
        if inputparam == "it":
            if np.sign(self.it.U[0])!=-np.sign(omega):
                self.it.U *= -1
                self.it.phi *= -1
                self.it.alpha = np.arctan(self.it.vx/self.it.vtheta)
                self.it.wtheta = self.it.vtheta - self.it.U
                self.it.w = np.sqrt(self.it.wx**2 + self.it.wtheta**2)
                self.it.beta = np.abs(np.arctan(self.it.wx/self.it.wtheta)) + self.gp.angpitch*np.pi/180
            Um = self.it.U[len(self.it.U)//2]
            phi = self.it.phi[len(self.it.phi)//2]
            dq_aux = np.abs(self.it.dq)
        else:
            # Tangential velocity calculation.
            self.it.U = -np.array([omega*_ for _ in self.gp.r])
            self.ot.U = copy.deepcopy(self.it.U)
            # Midspan tangential velocity instantiation.
            Um = self.it.U[len(self.it.U)//2]
            # Axial velocity, midspan flow parameter and flow coefficient calculations depending on the provided input parameter.
            if inputparam=="phi":
                # If the input parameter is a flow coefficient, then:
                # check, first, if the input variable is an array or nself.ot.        
                if isinstance(phi_q_v, np.ndarray):
                    # If the is, then set the axial velocity to the product of the input parameter and the local (element-wise) tangential velocity.
                    self.it.vx = np.array([_[0]*_[1] for _ in zip(phi_q_v, self.it.U)])
                else:
                    # If it isn't, then the axial velocity of the inlet velocity triangle is obtained by multiplying phi and Um, assuming a constant axial velocity radially.
                    self.it.vx = np.array([phi_q_v*Um for _ in self.gp.r])                    
                # the midspan flow parameter coincides with the input value.
                phi = phi_q_v
                # The differential flow-rate is calculated from its definition.
                self.it.dq = 2*np.pi*self.gp.r*self.it.vx
                # The flow-rate is calculated by integration.
                self.it.q = simpson(self.it.dq, dx=self.gp.dr)
            elif inputparam=="q":
                # If the input parameter is a flow rate value, then:
                # check, first, if the input variable is an array or nself.ot.
                if isinstance(phi_q_v, np.ndarray):
                    # If it is, then set the axial velocity from its definition of differential flow-rate.
                    self.it.vx = phi_q_v/(2*np.pi*self.gp.r)
                    # Set the differential flow-rate to the input variable.
                    self.it.dq = phi_q_v
                    # Set the flow-rate to the integral value of the differential flow-rate.
                    self.it.q = simpson(self.it.dq, dx=self.gp.dr)
                else:
                    # If it isn't, then set the axial velocity from its definition of integral flow-rate, assuming constant axial velocity of the inlet velocity triangle.
                    self.it.vx = np.array([phi_q_v/(np.pi*(self.gp.rtip**2 - self.gp.rhub**2)) for _ in self.gp.r])
                    # Set the integral flow-rate to the input variable.
                    self.it.q = phi_q_v
                    # Set the differential flow-rate from its definition.
                    self.it.dq = 2*np.pi*self.gp.r*self.it.vx
                # The midspan flow parameter is obtained by dividing the axial velocity with the midspan tangential velocity.
                phi = self.it.vx[len(self.it.vx)//2]/Um
            else:
                # Check, first, if the input variable is an array or nself.ot.
                if isinstance(phi_q_v, np.ndarray):
                    # If it is, then set the axial velocity component of the inlet velocity triangle to the input variable.
                    self.it.vx = phi_q_v
                else:
                    # If it isn't, then set the axial velocity component of the inlet velocity triangle to a constant value equal to the input parameter.
                    self.it.vx = np.array([phi_q_v for _ in self.gp.r])
                # The midspan flow paramter is obtained by dividing the axial velocity with the midspan tangential velocity.
                phi = self.it.vx[len(self.it.vx)//2]/Um    
                # The differential flow-rate is calculated from its definition.
                self.it.dq = 2*np.pi*self.gp.r*self.it.vx
                # The flow-rate is calculated by integration.
                self.it.q = simpson(self.it.dq, dx=self.gp.dr)
            # Setting the non-dimensional flow-coefficient at the input triangle.
            self.it.Phi = self.it.q/(np.abs(omega)*(2*self.gp.rtip)**3)
            # The alpha angle (between the absolute and tangential velocities) is set to pi/2 by assumption.
            self.it.alpha = np.array([np.pi/2 for i in range(len(self.it.alpha))])
            # The magnitude of the absolute velocity is obtained by pitagorean application.
            self.it.v = np.sqrt(self.it.vx**2 + self.it.vtheta**2)
            # The axial component of the relative velocity is set to the axial component of the absolute velocity.
            self.it.wx = np.array([_ for _ in self.it.vx])
            # The tangential component of the relative velocity is set to (-) the tangential velocity, from the vectorial identity.
            self.it.wtheta = np.array([-_ for _ in self.it.U])
            # The magnitude of the relative velocity is obtained by pitagorean application.
            self.it.w = np.sqrt(self.it.wx**2 + self.it.wtheta**2)
            # The beta1 angle (between the relative and tangential velocities) is obtained from trigonometric relations.
            self.it.beta = np.arctan(self.it.wx/self.it.wtheta) + self.gp.angpitch*np.pi/180
            # The radially-varying flow coefficient is obtained from its definition.
            self.it.phi = self.it.vx/self.it.U
            # Expressing an auxiliar differential flow-rate as the absolute value of the flow-rate (necessary for loss calculations).
            dq_aux = np.abs(self.it.dq)

        #####
        ##### SETTING REFERENTIAL MAGNITUDE
        #####
        # Declaring 'vx', 'v' and w' as referential velocity magnitudes at the disc and setting them to 0.
        vx = 0
        v = 0
        w = 0
        # Conditional for setting the referential velocity magnitudes according to the value of 'refvel'.
        if refvel == "it":
            # If 'refvel' is 'it', then the referential magnitude is set equal to the inlet velocity triangle's relative magnitude.
            vx = self.it.vx
            v = self.it.v
            w = self.it.w
        else:
            # If 'refvel' is 'ave', then the referential velocity magnitudes are set equal to the velocity magnitudes at the disk.
            vx = 0.5*(self.ot.vx + self.it.vx)
            v = np.sqrt(vx**2 + self.it.vtheta**2)
            w = np.sqrt(vx**2 + self.it.wtheta**2)

        #####
        ##### REYNOLDS CALCULATION AND BILINEAR INTERPOLATION FOR GETTING CL AND CD VALUES
        #####
        # The Reynolds number is obtained from its definition, Re=p·W·c/u, where p is the density and nu the viscosity.        
        self.coefs.Re = self.fp.rho*self.gp.chord*w/self.fp.nu        
        # Loop running over the inlet velocity triangle's relative angle values.
        for e, _ in enumerate(self.it.beta):
            _ *= 180/np.pi
            # Getting the lift and drag coefficient polars corresponding to current blade element's airfoil geometry.
            polar_cl = polars[self.gp.airfoil[e] + "CL"]
            polar_cd = polars[self.gp.airfoil[e] + "CD"]
            # Getting the Reynolds values for which the lift and drag coefficients have been obtained.
            Re_values = list(polar_cl.keys())[1:]            
            # Getting the index that determines between which tested Reynolds values falls the current Reynolds, and retrieving those extreme Reynolds values for interpolation.
            if all([self.coefs.Re[e] - __ > 0 for __ in Re_values]):
                minindex = len(Re_values) - 2
            else:
                minindex = [self.coefs.Re[e] - __ < 0 for __ in Re_values].index(True) - 1
            Remin = Re_values[minindex]
            Remax = Re_values[minindex + 1]
            # Getting the lift and drag datasets that correspond to the minimum and maximum Reynolds values.
            lifts = [polar_cl[Remin], polar_cl[Remax]]            
            drags = [polar_cd[Remin], polar_cd[Remax]]
            # Retrieving the angles-of-attack for which the coefficients have been obtained.
            alpha = polar_cl['angle'] 
            alpha_0ind = list(alpha).index(0)
            diffalpha = list(alpha - _)                
            floorind = [__ > 0 for __ in diffalpha].index(True) - 1        
            ceilind = [__ > 0 for __ in diffalpha].index(True)       
            # Maximum and minimum lift values are determined, given the current angle-of-attack, so that the double interpolation scheme is applied.        
            clminmin, clminmax = lifts[0][floorind], lifts[0][ceilind]
            clmaxmin, clmaxmax = lifts[1][floorind], lifts[1][ceilind]
            # Calling the double_interp() method with current angle-of-attack and Reynolds number, thus determining the lift value.
            self.coefs.cl[e] = double_interp(np.floor(_), Remin, np.ceil(_), Remax, _, self.coefs.Re[e], clminmin, clmaxmin, clminmax, clmaxmax)            
            # Maximum and minimum drag values are determined, given the current angle-of-attack, so that the double interpolation scheme is applied.
            cdminmin, cdminmax = drags[0][floorind], drags[0][ceilind]
            cdmaxmin, cdmaxmax = drags[1][floorind], drags[1][ceilind]
            # Calling the double_interp() method with current angle-of-attack and Reynolds number, thus determining the drag value.
            self.coefs.cd[e] = double_interp(np.floor(_), Remin, np.ceil(_), Remax, _, self.coefs.Re[e], cdminmin, cdmaxmin, cdminmax, cdmaxmax)        
            
            #####
            ##### LIFT-COEFFICIENT-BASED BLADE-ELEMENT STATE DETERMINATION
            #####
            # Instantiating 'minliftangles' and 'maxliftangles' lists for storing the angles at which the minimum and maximum lift coefficients occur for the extreme Reynolds cases.
            minliftangles = list()
            maxliftangles = list()
            # Loop for retrieving minimum and maximum lift values and their corresponding angles within the polar.
            for f, __ in enumerate(lifts):
                # The 'difflifts' stores the difference between adjacent absolute-valued lifts.
                difflifts = list(np.diff(np.abs(lifts[f])))
                # The 'minliftangles' and 'maxliftangles' are set to those angles at which the difference in absolute-valued lifts falls below 0, which means that a decay on the lifts begins.
                minliftangles.append(alpha[alpha_0ind - [-___ < 0 for ___ in difflifts[:alpha_0ind][::-1]].index(True) - 1])
                maxliftangles.append(alpha[alpha_0ind + [___ < 0 for ___ in difflifts[alpha_0ind:]].index(True) - 1])
            # The computed angles are sorted.
            minliftangles = sorted(minliftangles)
            maxliftangles = sorted(maxliftangles)
            # Conditional on which the lift-coefficient-based blade-element state is determined.
            # If the current lift coefficient is >0:
            if self.coefs.cl[e] > 0:
                # If the current blade-element's angle-of-attack lies below the computed maximum lift angles, then the state is set to 'below_max'.
                if all([_ < __ for __ in maxliftangles]):
                    self.coefs.clstate[e] = "below_max"
                # Else, if the current blade-element's angle-of-attack lies above the computed maximum lift angles, then the state is set to 'above_max'.
                elif all([_ > __ for __ in maxliftangles]):
                    self.coefs.clstate[e] = "above_max"
                # Otherwise, it is necessary to perform an interpolation for knowing if the current angle-of-attack lies above or below the maximum lift corresponding to the current Reynolds number.
                else:
                    # The 'cl_comp_list' is a list for storing the lift coefficient values that are to be employed for comparison.
                    cl_comp_list = list()
                    # If the computed maximum lift coefficients do not coincide, then there are a number of values at which the comparison is to be made.
                    if maxliftangles[0] != maxliftangles[1]:
                        # Setting 'alpha_comp_list' to those angles-of-attack.
                        alpha_comp_list = alpha[list(alpha).index(maxliftangles[0]):list(alpha).index(maxliftangles[1])]
                    # Otherwise, if the computed maximum lift coefficients coincide, then there is just one value at which the comparison is to be made.
                    else:
                        # Setting 'alpha_comp_list' to the corresponding angle-of-attack of comparison.
                        alpha_comp_list.append(maxliftangles[0])
                    # Loop for performing the interpolation at each angle-of-attack of comparison, and appending the resultant values to 'cl_comp_list'.
                    for a in alpha_comp_list:
                            cl_comp_list.append(double_interp(a, Remin, a, Remax, a, self.coefs.Re[e],
                                                              lifts[0][list(alpha).index(maxliftangles[0])],
                                                              lifts[1][list(alpha).index(maxliftangles[0])],
                                                              lifts[0][list(alpha).index(maxliftangles[1])],
                                                              lifts[1][list(alpha).index(maxliftangles[1])]))
                    # Computing the angle-of-attack at which the maximum lift-coefficient value occurs for the current Reynolds number.
                    maxliftalpha = alpha_comp_list[list(cl_comp_list).index(max(cl_comp_list))]
                    # Setting element's lift-coefficient-based state according to the computed angle-of-attack at which the maximum lift coefficient occurs.
                    if _ < maxliftalpha:
                        self.coefs.clstate[e] = "below_max"
                    else:
                        self.coefs.clstate[e] = "above_max"
            # If the current lift coefficient is <0:
            else:
                # If the current blade-element's angle-of-attack lies above the computed minimium lift angles, then the state is set to 'below_max'.
                if all([_ > __ for __ in minliftangles]):
                    self.coefs.clstate[e] = "below_max"
                # Else, if the current blade-element's angle-of-attack lies below the computed minimum lift angles, then the state is set to 'above_max'.
                elif all([_ < __ for __ in minliftangles]):
                    self.coefs.clstate[e] = "above_max"
                # Otherwise, it is necessary to perform an interpolation for knowing if the current angle-of-attack lies above or below the minimum lift corresponding to the current Reynolds number.
                else:
                    # The 'cl_comp_list' is a list for storing the lift coefficient values that are to be employed for comparison.
                    cl_comp_list = list()
                    # If the computed minimum lift coefficients do not coincide, then there are a number of values at which the comparison is to be made.                    
                    if minliftangles[0] != minliftangles[1]:
                        # Setting 'alpha_comp_list' to those angles-of-attack.
                        alpha_comp_list = alpha[list(alpha).index(minliftangles[0]):list(alpha).index(minliftangles[1])]
                    # Otherwise, if the computed minimum lift coefficients coincide, then there is just one value at which the comparison is to be made.
                    else:
                        # Setting 'alpha_comp_list' to the corresponding angle-of-attack of comparison.
                        alpha_comp_list.append(minliftangles[0])
                    # Loop for performing the interpolation at each angle-of-attack of comparison, and appending the resultant values to 'cl_comp_list'.
                    for a in alpha_comp_list:
                            cl_comp_list.append(double_interp(a, Remin, a, Remax, a, self.coefs.Re[e],
                                                              lifts[0][list(alpha).index(minliftangles[0])],
                                                              lifts[1][list(alpha).index(minliftangles[0])],
                                                              lifts[0][list(alpha).index(minliftangles[1])],
                                                              lifts[1][list(alpha).index(minliftangles[1])]))
                    # Computing the angle-of-attack at which the minimum lift-coefficient value occurs for the current Reynolds number.
                    minliftalpha = alpha_comp_list[list(cl_comp_list).index(min(cl_comp_list))]
                    # Setting element's lift-coefficient-based state according to the computed angle-of-attack at which the minimum lift coefficient occurs.
                    if _ > minliftalpha:
                        self.coefs.clstate[e] = "below_max"
                    else:
                        self.coefs.clstate[e] = "above_max"
        # Performing the curve-fitting if the value of the flow parameter is not 0 (null lift value).
        if phi != 0:
            # Lift-fitting.
            self.coefs.cl = dynamic_fitting(self.it.beta*180/np.pi, self.coefs.cl, 2, resmin=1e-2)
            # Drag-fitting.
            self.coefs.cd = dynamic_fitting(self.it.beta*180/np.pi, self.coefs.cd, 2, resmin=1e-2)     

        #####
        ##### COMPUTATION OF DIFFERENTIAL LOADS, INDUCED DRAG, THRUST AND TANGENTIAL-FORCE COEFFICIENTS
        #####
        # Differential lift computation.
        self.coefs.dl = self.coefs.cl*0.5*self.fp.rho*(w**2)*self.gp.dA
        # Differential drag computation.
        self.coefs.dd = self.coefs.cd*0.5*self.fp.rho*(w**2)*self.gp.dA
        # Induced drag coefficients from its formulation.
        self.coefs.cdtc = self.coefs.cd + 0.7*np.abs(self.coefs.cl)*self.gp.tip_c/(self.gp.AR*self.gp.pitch)
        # Differential induced drag load computation.
        self.coefs.ddtc = self.coefs.cdtc*0.5*self.fp.rho*(w**2)*self.gp.dA
        # Projection for obtaining axial coefficient.
        self.coefs.cx = self.coefs.cl*np.cos(self.it.beta) + self.coefs.cdtc*np.sin(self.it.beta)
        # Differential axial load computation.
        self.coefs.dx = self.coefs.cx*0.5*self.fp.rho*(w**2)*self.gp.dA
        # Projection for obtaining tangential coefficient.
        self.coefs.ctheta = -self.coefs.cl*np.sin(self.it.beta) + self.coefs.cdtc*np.cos(self.it.beta)
        # Differential tangential load computation.
        self.coefs.dtheta = self.coefs.ctheta*0.5*self.fp.rho*(w**2)*self.gp.dA

        #####
        ##### COMPUTATION OF INTERFERENCE-CORRECTED (SOLIDITY-BASED SEMI-EMPIRICAL CORRECTION) THRUST AND TANGENTIAL-FORCE COEFFICIENTS
        #####
        # Correction of axial coefficient.
        self.coefs.cxCIF = self.coefs.cx/(1 - self.gp.sigma**2)
        # Correction of axial load.
        self.coefs.dxCIF = self.coefs.dx/(1 - self.gp.sigma**2)
        # Correction of tangential coefficient.
        self.coefs.cthetaCIF = self.coefs.ctheta/(1 - self.gp.sigma**2)       
        # Correction of tangential load.
        self.coefs.dthetaCIF = self.coefs.dtheta/(1 - self.gp.sigma**2)

        #####
        ##### DOWNSTREAM VELOCITY-TRIANGLE CALCULATION
        #####
        # Conditional for passing in case the beta1 angle is null (no flow).
        if 0 in self.it.beta:
            pass
        else:
            # Calculation of the absolute angle of the outlet velocity triangle.
            self.ot.alpha = np.pi/2 - np.arctan((1/np.tan(self.it.alpha))-(self.gp.Z*self.gp.chord*self.coefs.cthetaCIF)/(4*np.pi*self.gp.r*np.sin(np.abs(self.it.beta))**2))        
            # The downstream absolute axial velocity component is set equal to the upstream one, by modelization assumption.
            self.ot.vx = copy.deepcopy(self.it.vx)
            # The tangential component is obtained via trigonometrical relation.        
            self.ot.vtheta = np.sign(omega)*np.abs(self.ot.vx)/np.tan(self.ot.alpha)
            # Instantiating the 'vx2' variable as the square of the axial component of the outlet velocity triangle.
            vx2 = self.it.vx**2
            
            #####
            ##### RADIAL-EQUILIBRIUM CONDITION ENFORCEMENT
            #####
            # Instantiating the 'gamma' variable, which represents the element-wise circulation along the blade.
            gamma = 2*np.pi*self.gp.r*self.ot.vtheta       
            # Instantiating the 'dvx2_dr' variable, which represents the derivative of the square of the axial velocity component with respect to the radial parameter.
            dvx2_dr = (-1/(2*np.pi*self.gp.r)**2)*(gamma**2)        
            # Computing the radial values of the square of the axial velocity component of the outlet velocity triangle from the obtained differences.
            # The loop starts from the second blade element, meaning that the velocity value of the first element is kept unchanged on a first guess/approximation.
            for e, _ in enumerate(self.gp.r[1:]):
                e += 1
                vx2[e] = vx2[e-1] + dvx2_dr[e]*self.gp.dr                   
            # Instantiating the 'diff_q' variable and setting its value to 1, representing the threshold for fulfilling the continuity condition.
            diff_q = 1
            # Instantiating the 'diff_prev' variable and setting its value to 2 (so that, in the first step, 'diff_q_prev > diff_q'). 
            diff_q_prev = 2
            # Instantiating the 'sign' variable and setting its value to 1.
            sign = 1
            # Instantiating the 'reverse' variable and setting its value to 'False'.
            reverse = False
            # Continuity-equation enforcement.
            while diff_q > self.it.q*0.1:
                # If there are no negative square elements, then the axial velocity component of the outlet velocity triangle is set equal to the square-root of the obtained distribution of the square velocity.
                if min(vx2) >= 0:
                    self.ot.vx = np.sqrt(vx2)
                # If there are negative square elements, those are computed by making the negative value absolute and making the overall velocity component change its sign.
                else:
                    break
                # The 'diff' and 'diffaux' variables compute the difference between a current-step flow-rate and the flow-rate provided by the inlet velocity triangle, and the previous-step flow-rate and idem, respecively (check the while loop below).                
                diff_q = np.abs(np.abs(self.it.q) - np.abs(simpson(2*np.pi*self.gp.r*self.ot.vx, dx=self.gp.dr)))
                # Condition for checking whether the current step's continuity condition fulfillment is further from the previous step's one; if it is, then continuity divergence is taking place. The sign is reversed, and the 'reverse' variable set to 'True'.
                if diff_q_prev - diff_q < 0 and not reverse:
                    sign *= -1
                    reverse = True
                # Setting 'diff_q_prev' to 'diff_q' for the next step's continuity condition checking.
                diff_q_prev = diff_q
                # Varying the square of the axial velocity component differentially.
                vx2 += sign*1e-2
            # Casting the outlet velocity triangle's axial velocity components to match their original signs.
            self.ot.vx *= np.sign(self.it.vx[0])
            # Recomputing the outlet velocity triangle's absolute angle.
            self.ot.alpha = np.arctan(self.ot.vx/self.ot.vtheta)
            
            #####
            ##### RESUMING THE CALCULATIONS OF THE OUTLET VELOCITY TRIANGLE
            #####
            # Calculating the flow-rate at the outlet velocity triangle, which should coincide with the inlet flow-rate due to continuity enforcement.
            self.ot.q = simpson(2*np.pi*self.gp.r*self.ot.vx, dx=self.gp.dr)
            # Setting the non-dimensional flow-coefficient at the output triangle.
            self.ot.Phi = self.ot.q/(omega*(2*self.gp.rtip)**3)
            # The pitagorean theorem is used for calculating the absolute velocity magnitude of the outlet velocity triangle.
            self.ot.v = np.sqrt(self.ot.vx**2 + self.ot.vtheta**2)        
            # The axial velocity component of the relative velocity of the outlet velocity triangle is set equal to the absolute axial component, by trigonometric relations.
            self.ot.wx = self.ot.vx
            # The tangential component of the relative velocity of the outlet velocity triangle is obtained from vector identity.
            self.ot.wtheta = self.ot.vtheta - self.ot.U     
            # The pitagorean theorem is used for calculating the relative velocity magnitude of the outlet velocity triangle.
            self.ot.w = np.sqrt(self.ot.wx**2 + self.ot.wtheta**2)       
            # The relative angle of the outlet velocity triangle is obtained from trigonometric relations.
            self.ot.beta = np.arctan(self.ot.wx/self.ot.wtheta)

        #####
        ##### CALCULATION OF DIFFERENTIAL TORQUES, POWERS, PRESSURE DROPS AND EFFICIENCIES
        #####
        # Calculation of the differential torque.
        self.eout.difftau = -0.5*self.fp.rho*self.gp.chord*self.gp.Z*self.coefs.cthetaCIF*(w**2)*self.gp.r    
        # Calculation of integrated torque.
        self.eout.tau = simpson(self.eout.difftau, dx=self.gp.dr)    
        # Calculation of torque coefficient.
        if 0 not in vx and 0 not in v:
            self.coefs.ctau = self.eout.tau/(0.5*self.fp.rho*(v[len(v)//2]**2)*self.gp.chord*self.gp.rtip*(1 - self.gp.hub_to_tip_ratio)*self.gp.Z*vx[len(vx)//2])
        # Calculation of the differential power, multiplying the torque with the angular velocity.
        self.eout.diffpower = self.eout.difftau*np.abs(omega)
        # Calculation of integrated power.
        self.eout.power = simpson(self.eout.diffpower, dx=self.gp.dr)
        # Computing the non-dimensional output power coefficient.
        self.eout.Pi = self.eout.power/(self.fp.rho*(np.abs(omega)**3)*(2*self.gp.rtip)**5)
        # Conditional for setting the radially-varying powers, pressure drops and efficiencies to 0 in case beta1=0.
        if 0 in self.it.beta:
            pass
        else:
            # Calculation of differential static-to-static-pressure-drop.
            self.eout.dp_st_to_st = (self.fp.rho*self.gp.chord*self.gp.Z*self.coefs.cxCIF*(w**2))/(4*np.pi*self.gp.r)       
            # Calculation of static-to-static-pressure-drop.
            self.eout.p_st_to_st = simpson(2*np.pi*self.gp.r*self.eout.dp_st_to_st, dx=self.gp.dr)/simpson(2*np.pi*self.gp.r, dx=self.gp.dr)
            # Calculation of input coefficient.
            if 0 not in vx and 0 not in v:
                self.coefs.cinput = self.eout.p_st_to_st*self.it.q/(0.5*self.fp.rho*(v[len(v)//2]**2)*self.gp.chord*self.gp.rtip*(1 - self.gp.hub_to_tip_ratio)*self.gp.Z*vx[len(vx)//2])
            self.eout.dp_tot_to_st = -(np.abs(self.eout.dp_st_to_st) + 0.5*self.fp.rho*(vx**2))
            # Calculation of total-to-static pressure-drop.
            self.eout.p_tot_to_st = simpson(2*np.pi*self.gp.r*self.eout.dp_tot_to_st, dx=self.gp.dr)/simpson(2*np.pi*self.gp.r, dx=self.gp.dr)        
            # Calculation of differential total-to-total pressure-drop.
            self.eout.dp_tot_to_tot = -(np.abs(self.eout.dp_tot_to_st) - np.abs(0.5*self.fp.rho*self.ot.v**2))
            # Calculation of total-to-total pressure-drop.
            self.eout.p_tot_to_tot = simpson(2*np.pi*self.gp.r*self.eout.dp_tot_to_tot, dx=self.gp.dr)/simpson(2*np.pi*self.gp.r, dx=self.gp.dr)
            # Computing the non-dimensional power coefficient.
            self.eout.Psi = np.abs(self.eout.p_tot_to_st/(self.fp.rho*(np.abs(omega)**2)*(2*self.gp.rtip)**2))       
            # Calculation of differential loss power due to viscous effects.       
            self.eout.dpvisc = 0.5*self.fp.rho*self.gp.chord*self.gp.Z*self.coefs.cdtc*(w**3)    
            # Calculation of loss power due to viscous effects.
            self.eout.pvisc = simpson(self.eout.dpvisc, dx=self.gp.dr)
            # Calculation of differential loss power due to outlet kinetic velocity.
            self.eout.dpk = 0.5*self.fp.rho*(self.ot.v**2)*dq_aux
            # Calculation of loss power due to outlet kinetic energy.
            self.eout.pk = simpson(self.eout.dpk, dx=self.gp.dr)
            # Calculation of input power.
            dpin = self.eout.diffpower + self.eout.dpvisc + self.eout.dpk
            # Calculation of differential power.
            self.eout.deff = self.eout.diffpower/dpin
            # Calculation of efficiency.                
            self.eout.eff = simpson(self.eout.diffpower, dx=self.gp.dr)/simpson(dpin, dx=self.gp.dr)
            # Computing the efficiency derived from non-dimensional parameters.
            self.eout.Eta = np.abs(self.eout.Pi/(self.eout.Psi*self.ot.Phi)) if np.abs(self.eout.Pi/(self.eout.Psi*self.ot.Phi)) < 1 else 0
            
        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------------RETURN------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#             
            
        # Return statement.
        return

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------BEM_wells_ad() METHOD----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    def BEM_wells_ad(self,
                     phi_q_v: Union[list, np.ndarray, float],
                     omega: float,
                     polars: dict,
                     inputparam: str="phi",
                     refvel: str='it',
                     maxiter: int=100) -> None:
        '''Applies the BEM coupled with the actuator disc model to a Wells turbine stage.

        - **parameters**, **returns**, **return types**

        :param ts: instance of the turbine_stage class to which the BEM is to be applied.
        :param phi_q_v: input parameter, either a flow coefficient (phi) a flow-rate value (q) or
        an axial velocity value (vx). It may be either an array-like structure, specifying phi, q
        or v for each blade element, or a single value, in which the provided value is set equal
        to all the blade elements.
        :param omega: rotational speed of the turbine [rad/s].
        :param polars: a dictionary containing the polars of the element airfoils.
        :param inputparam: string representing the type of input provided; it is "phi" for indicating
        a flow coefficient, "q" for a flow-rate, "vx" for an axial velocity or "it" for a pre-established
        inlet velocity triangle. Default is "phi".
        :param refvel: string representing the referential velocity employed for calculating loads
        and velocity-dependent outputs; it is "it" for setting the referential velocity equal to the
        relative velocity magnitude of the inlet velocity triangle, or 'ave' for setting it equal to
        the relative velocity magnitude based on the actuator disk value (the axial component being
        the average of downstream and upstream velocities). Default is "it".
        :param maxiter: integer representing the maximum number of iterations to perform. Default is 100.
        '''

        #--------------------------------------------------------------------------------------------------------------------#
        #---------------------------------------------------------BODY-------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#        
        
        # Calling the BEM_wells_noad() method in the first step. The 'refvel' magnitude is set to 'it',
        # as the referential magnitude is taken from inlet velocity triangle in the first step.
        self.BEM_wells_noad(phi_q_v, omega, polars, inputparam, refvel='it')
        # Instantiating the 'diff_powout' variable and setting its value to 1; it is used for ensuring energy convservation convergency.
        diff_powout = 1
        # Instantiating the 'powout' variable and setting it equal to the total-to-total pressure drop.
        powout = self.eout.power + self.eout.pvisc + self.eout.pk
        # While loop for ensuring energy consrevation.
        ind = 0
        while diff_powout > 1e-4:       
            # Calling the BEM_wells_noad() method with the 'refvel' parameter set to 'ave'.
            self.BEM_wells_noad(phi_q_v, omega, polars, inputparam="it", refvel="ave")
            # Recomputing the 'diff_powout' value.
            diff_powout = np.abs(powout - (self.eout.power + self.eout.pvisc + self.eout.pk))
            # Reassigning the 'powout' value to current step's total-to-total pressure drop.
            powout = self.eout.power + self.eout.pvisc + self.eout.pk
            ind += 1
            if ind > maxiter:
                break
                
        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------------RETURN------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
                
        # Return statement.
        return
    
#--------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------------CLASS turbine--------------------------------------------------#             
#--------------------------------------------------------------------------------------------------------------------#
class turbine:
    '''Sets a turbine.'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------__init__() METHOD-------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#  
    def __init__(self) -> None:
        '''Initializes a given instance of a class containing a turbine.'''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------DECLARING ATTRIBUTES-----------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Number of stages.
        self._numstages = 0
        # Types of stages.
        self._typestages = list()
        # Stages list.
        self._stages = list()
        # Global dimensional torque.
        self._torque = list()
        # Dimensoinal torques.
        self._torque_stages = list()
        # Global dimensional power.
        self._power = list()
        # Dimensional powers.
        self._power_stages = list()
        # Global static-to-static pressure loss.
        self._pst = list()
        # Static-to-static pressure losses.
        self._pst_stages = list()
        # Global total-to-total pressure loss.
        self._ptt = list()
        # Total-to-total pressure losses.
        self._ptt_stages = list()
        # Dimensionless flow-rates.
        self._Phi = list()
        # Global dimensionless powers.
        self._Pi = list()
        # Dimensionless powers correspodning to each stage.
        self._Pi_stages = list()
        # Global dimensionless pressure losses.
        self._Psi = list()
        # Dimensionless pressure losses correspodning to each stage.
        self._Psi_stages = list()
        # Global efficiency.
        self._Eff = list()
        # Efficiencies corresponding to each stage.
        self._Eff_stages = list()
        
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------DECLARING PROPERTIES-----------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------# 
    
    # Declaration of get property numstages for getting the number of stages.
    @property
    def numstages(self) -> int:
        return self._numstages
    # Declaration of get property typestages for getting the list of the types of stages.
    @property
    def typestages(self) -> list:
        return self._typestages
    # Declaration of get property stages for getting the list of stages.
    @property
    def stages(self) -> list:
        return self._stages
    # Declaration of get property torque for getting the glboal dimensional torque.
    @property
    def torque(self) -> list:
        return self._torque
    # Declaration of get property torque_stages for getting the dimensional torques corresponding to each stage.
    @property
    def torque_stages(self) -> list:
        return self._torque_stages
    # Declaration of get property power for getting the glboal dimensional torque.
    @property
    def power(self) -> list:
        return self._power
    # Declaration of get property power_stages for getting the dimensional powers corresponding to each stage.
    @property
    def power_stages(self) -> list:
        return self._power_stages    
    # Declaration of get property pst for getting the glboal dimensional static-to-static pressure loss.
    @property
    def pst(self) -> list:
        return self._pst
    # Declaration of get property pst_stages for getting the dimensional static-to-static pressure losses corresponding to each stage.
    @property
    def pst_stages(self) -> list:
        return self._pst_stages        
    # Declaration of get property ptt for getting the glboal dimensional total-to-total pressure loss.
    @property
    def ptt(self) -> list:
        return self._ptt
    # Declaration of get property ptt_stages for getting the dimensional total-to-total pressure losses corresponding to each stage.
    @property
    def ptt_stages(self) -> list:
        return self._ptt_stages            
    # Declaration of get property Phi for getting the list of dimensionless flow-rates.
    @property
    def Phi(self) -> list:
        return self._Phi
    # Declaration of get property Pi for getting the list of global dimensionless powers.
    @property
    def Pi(self) -> list:
        return self._Pi
    # Declaration of get property Pi_stages for getting the list of dimensionless powers corresponding to each stage.
    @property
    def Pi_stages(self) -> list:
        return self._Pi_stages
    # Declaration of get property Psi for getting the list of global dimensionless pressure losses.
    @property
    def Psi(self) -> list:
        return self._Psi
    # Declaration of get property Psi_stages for getting the list of dimensionless pressure losses corresponding to each stage.
    @property
    def Psi_stages(self) -> list:
        return self._Psi_stages
    # Declaration of get property Eff for getting the list of global efficiencies.
    @property
    def Eff(self) -> list:
        return self._Eff
    # Declaration of get property Eff_stages for getting the list of efficiencies corresponding to each stage.
    @property
    def Eff_stages(self) -> list:
        return self._Eff_stages       
    
    # Declaration of set property torque for global dimensional torque list.
    @torque.setter 
    def torque(self, value: list) -> None:    
        self._torque = value
        return    
    # Declaration of set property torque_stages for dimensional torques list corresponding to each stage.
    @torque_stages.setter 
    def torque_stages(self, value: list) -> None:    
        self._torque_stages = value
        return
    # Declaration of set property power for global dimensional power list.
    @power.setter 
    def power(self, value: list) -> None:    
        self._power = value
        return    
    # Declaration of set property power_stages for dimensional powers list corresponding to each stage.
    @power_stages.setter 
    def power_stages(self, value: list) -> None:    
        self._power_stages = value
        return
    # Declaration of set property power pst global dimensional static-to-static pressure loss list.
    @pst.setter 
    def pst(self, value: list) -> None:    
        self._pst = value
        return    
    # Declaration of set property pst_stages for dimensional static-to-static pressure loss list corresponding to each stage.
    @pst_stages.setter 
    def pst_stages(self, value: list) -> None:    
        self._pst_stages = value
        return
    # Declaration of set property power ptt global dimensional total-to-total pressure loss list.
    @ptt.setter 
    def ptt(self, value: list) -> None:    
        self._ptt = value
        return    
    # Declaration of set property ptt_stages for dimensional total-to-total pressure loss list corresponding to each stage.
    @ptt_stages.setter 
    def ptt_stages(self, value: list) -> None:    
        self._ptt_stages = value
        return    
    # Declaration of set property Phi for dimensionless flow-rate list.
    @Phi.setter 
    def Phi(self, value: list) -> None:    
        self._Phi = value
        return
    # Declaration of set property Pi for global dimensionless power list.
    @Pi.setter 
    def Pi(self, value: list) -> None:    
        self._Pi = value
        return
    # Declaration of set property Pi_stages for dimensionless power list corresponding to each stage.
    @Pi_stages.setter 
    def Pi_stages(self, value: list) -> None:    
        self._Pi_stages = value
        return
    # Declaration of set property Psi for global dimensionless pressure loss list.
    @Psi.setter 
    def Psi(self, value: list) -> None:    
        self._Psi = value
        return
    # Declaration of set property Psi_stages for dimensionless pressure loss list corresponding to each stage.
    @Psi_stages.setter 
    def Psi_stages(self, value: list) -> None:    
        self._Psi_stages = value
        return    
    # Declaration of set property Eff for global efficiency list.
    @Eff.setter 
    def Eff(self, value: list) -> None:    
        self._Eff = value
        return
    # Declaration of set property Eff_stages for efficiency list corresponding to each stage.
    @Eff_stages.setter 
    def Eff_stages(self, value: list) -> None:    
        self._Eff_stages = value
        return
    
    #--------------------------------------------------------------------------------------------------------------------#
    #---------------------------------------------------DEFINING METHODS-------------------------------------------------#
    #--------------------------------------------------------------------------------------------------------------------#     
    
    #--------------------------------------------------------------------------------------------------------------------#
    #------------------------------------------------reset_props() METHOD------------------------------------------------# 
    #--------------------------------------------------------------------------------------------------------------------#
    def reset_props(self) -> None:
        '''Resets the output global and stage-wise properties to empty lists.
        
        Necessary whenever a turbine stage is added or removed, or the BEM method re-applied to
        a turbine instance.       
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------------BODY--------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#        
        
        for prop_root_to_reset in ["torque", "power", "pst", "ptt", "Phi", "Pi", "Psi", "Eff"]:
            props_to_reset = [_ for _ in dir(self) if prop_root_to_reset in _]
            for prop_to_reset in props_to_reset:
                setattr(self, prop_to_reset, list())
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------create_or_add_stage() METHOD--------------------------------------------# 
    #--------------------------------------------------------------------------------------------------------------------#
    def create_or_add_stage(self,
                 pos: int=-1,
                 mode: str='create',
                 ts: Union[None, turbine_stage]=None,
                 ttype: str='Wells',
                 omega: float=3600,                 
                 N: int=50,
                 rcas: float=0.25,
                 hub_to_tip_ratio: float=0.75,
                 chord: Union[list, np.ndarray, float]=0.117,
                 angpitch: Union[list, np.ndarray, float]=0,
                 airfoil: Union[list, np.ndarray, str]="NACA0015",
                 tip_percent: float=0.5,
                 Z: int=7,
                 p: float=101325,
                 T: float=288.15,
                 R: float=287.058,
                 nu: float=1.81e-5) -> None:
        '''Creates or adds a turbine stage on the turbine, provided the input parameters for the stage.                
        
        **parameters**:
            :param pos: position on which to create the turbine stage. Default is -1, which means
            that the turbine stage is appended to the existing ones (if no turbine stage is present,
            then the turbine stage is set in the first position). Set it to 0 to prepend a stage.
            :param cr_or_add: string representing if a turbine stage is to be created ('create'),
            or added ('add'). Default is 'create'.
            :param ts: a turbine_stage instance in case an existing turbine stage is to be added.
            Default is None.
            :param ttype: string specifying turbine type; either 'Wells' or 'Impulse'. Default is
            'Wells'.
            :param omega: rotational speed of turbine stage, in [RPM]. Default is 3600 [RPM].
            :param N: number of blade elements (discretization). Default is 50.
            :param rcas: casing radius. Default is 0.25.
            :param hub_to_tip_ratio: hub to tip ratio value. Default is 0.75.
            :param chord: chord value. Default is 0.117. It can be an element-wise array.
            :param angpitch: angular pitch value. Default is 0. It can be an element-wise array.
            :param airfoil: geometry of the airfoils of the blade. Default is NACA0015, meaning that
             the airfoil geometry is constant throughout the blade. It can be an element-wise array.
            :param tip_percent: tip gap value, in terms of chord percentage. Default is 0.5%.
            :param Z: number of blades. Default is 7.
            :param p: atmospheric pressure in [Pa]. Default value is 101325 [Pa].
            :param T: ambient temperature in [K]. Default value is 288.15 [K].
            :param R: gas constant in [J/(kg·K)]. Default value is 287.058 [J/(kg·K)]
            :param nu: gas viscosity in [kg/(m·s)]. Default value is 1.81e-5 [kg/(m·s)]                 
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Asserting if the provided input parameter 'mode' has a valid value.
        assert mode in ["create", "add"], "Please provide a valid mode input, either 'create' or 'add'."

        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------------BODY--------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#        
        
        # Emptying output property lists.
        self.reset_props()
        
        # Conditional for instantiating 'numlist' variable. Just done if the number of stages is > 1.
        if self._numstages != 0:            
            # If the 'pos' variable is not set to append the created turbine stage, then 'numlist' must equal the stages beginning from 'pos + 1' and ending in the number of stages.
            if pos != -1:
                numlist = [_ + 1 for _ in range(pos, self._numstages)]
            # Otherwise, the numlist is an empty list.
            else:
                numlist = list()
        
        # If the 'pos' variable is set to append the created turbine stage, then 'num' must equal the number of stages + 1.
        if pos == -1:
            num = self._numstages + 1
        # Otherwise, it is necessary to overright all the stages that are after the position at which the stage to be created is to be inserted.
        else:
            # Conditional for ensuring that the number of stages is equal or greater than 1.
            if self._numstages != 0:
                # The 'num' variable is set to the first entry of 'numlist' in this case.
                num = numlist[0]
                # Loop running over the reverse 'numlist'. It gets the turbine stages in reverse order, form the last one to the 'pos + 1' variable.
                for i in numlist[::-1]:                    
                    # Getting the turbine stage instance.
                    ts = getattr(self, "turbstage" + str(i))
                    # Setting a property named 'turbstage(i+1)', and setting it equal to the turbine stage instance.
                    setattr(self, "turbstage" + str(i + 1), ts)
                    # Deleting turbine stage instance and corresponding output properties.
                    delattr(self, "turbstage" + str(i))
                    # Deleting instance from '_stages' list.
                    del self._stages[-1]
                
        # Conditional for either creating a turbine stage instance or adding an existing one.
        if mode == 'create':
            # If 'mode' is set to 'create', creating turbine stage instance.
            ts = turbine_stage(ttype=ttype,
                               omega=omega,                 
                               N=N,
                               rcas=rcas,
                               hub_to_tip_ratio=hub_to_tip_ratio,
                               chord=chord,
                               angpitch=angpitch,
                               airfoil=airfoil,
                               tip_percent=tip_percent,
                               Z=Z,
                               p=p,
                               T=T,
                               R=R,
                               nu=nu)
        else:
            # Otherwise, checking wheter the input turbine stage instance is valid.
            if ts == None:
                # If it is not, raising a type error.
                raise TypeError('No turbine stage provided to add. Please provide an existing turbine stage')
            else:
                # Otherwise, setting the 'ts' variable equal to the input turbine stage value.
                ts = ts
        
        # Setting an attribute for the created turbine instance, and setting it to that instance.
        setattr(self, "turbstage" + str(num), ts)
        # Appending the created turbine instance to '_stages'.
        self._stages.append(ts)
        # Conditional for appending the deleted turbine instances to '_stages'.
        if pos == -1:
            # 'pass' command in case the 'pos' variable is set to append the the created turbine stage to the existing stages.
            pass
        elif self._numstages != 0 and pos != -1:
            # If 'numlist' is not an empty list, it means that the created stage has been introduced somewhere between the previous stages, so that those existing stages where removed from the list.
            if len(numlist) != 0:
                for i in numlist:
                    # Appending the removed turbine stages to the '_stages' list.
                    self._stages.append(getattr(self, "turbstage" + str(i + 1)))
            # Otherwise, just appending the removed turbine stage to the '_stages' list.
            else:
                self._stages.append(getattr(self, "turbstage" + str(num)))
        # Updating the number of stages.
        self._numstages += 1
        # Updating the type of stages.
        self._typestages.append(self._stages[-1].ttype)
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------------RETURN-------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#        
        
        # Return statement.
        return

    #--------------------------------------------------------------------------------------------------------------------#
    #-----------------------------------------------remove_stage() METHOD------------------------------------------------# 
    #--------------------------------------------------------------------------------------------------------------------#
    def remove_stage(self, ind: int) -> None:
        '''Removes an existing turbine stage from the turbine.
        
         **parameters**:
            :param ind: integer determining the index of the list of stages that is removed.
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Asserting that the number of stages is larger than 0; that is, that there are turbine stage instances on the current turbine instance.
        assert self._numstages != 0, "There are no turbine stage instances in the current turbine instance. Skipping removal action."
        # Getting the maximum turbine stage number.
        maxturbstage = int([_ for _ in dir(self) if "turbstage" in _][-1].split("turbstage")[-1])
        # Asserting that the provided index to remove falls within the existing number of turbine stages.
        assert 0 < ind < maxturbstage, "The provided index does not fall within the existing number of turbine stages. Provide an index value between 1 and " + str(maxturbstage) + "."
        
        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------------BODY--------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#         
        
        # Emptying output property lists.
        self.reset_props()
        
        # The 'numlist' equals the stages beginning from 'ind + 1' and ending in the number of stages.            
        numlist = [_ + 1 for _ in range(ind, self._numstages - 1)]
        
        # Removing the turbine stage from the stages.
        del self._stages[ind - 1]
        # Deleting referred encapsulated turbine instance and its corresponding property.
        delattr(self, "turbstage" + str(ind))
        # Updating the number of stages.
        self._numstages -= 1
        # Updating the type of stages.
        del self._typestages[ind - 1]
        
        # Loop over 'numlist' for updating the numbers of the turbine stage instances and their corresponding properties
        for i in numlist:
            # Getting the turbine stage instance.
            ts = getattr(self, "turbstage" + str(i))
            # Setting a property named 'turbstage(i-1)', and setting it equal to the turbine stage instance.
            setattr(self, "turbstage" + str(i - 1), ts)
            # Deleting referred encapsulated turbine instance and its corresponding property.
            delattr(self, "turbstage" + str(i))
            
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------------RETURN-------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#             
            
        # Return statement.
        return
    
    #--------------------------------------------------------------------------------------------------------------------#
    #----------------------------------------------------BEM() METHOD----------------------------------------------------# 
    #--------------------------------------------------------------------------------------------------------------------#
    def BEM(self,
            phi_q_vs: Union[list, np.ndarray, float],
            polars: dict,
            inpres: Union[list, np.ndarray, float]=None,
            intemp: Union[list, np.ndarray, float]=None,
            inputparam: str='vx',
            mode: str='3DAD',
            reset: bool=True) -> None:
        '''Applies the momentum element method to the turbine.
        
        **parameters**, **return**, **return type**:                    
            :param phi_q_v: input parameter list for which the BEM method is applied. They can
            be either flow coefficient values (phi), flow-rate values (q) or axial velocity
            values (vx).
            :param polars: dictionary of lift and drag coefficients at different Reynolds numbers for
            a set of distinct airfoils from which to interpolate the lift and drag coefficients
            upon each of the blade elements.
            :param inpress: input pressure list. Default is 'None', which means that a standard value
            is provided (101325 Pa).
            :param intemp: input temperature list. Default is 'None', which means that a standard value
            is provided (288.15 K).
            :param inputparam: string determining the type of input parameter. It can be either
            'phi' (stipulating flow coefficient), 'flowrate' (flow-rate) or 'vx'
            (axial velocity). Default is 'vx'.
            :param mode: string determining the model of BEM to be applied. It can be either
            '3DNOAD' (semi-3D BEM model without actuator disk) or '3DAD' (semi-3D BEM model with
            actuator disk). Default is '3DAD'.  
            :param reset: resets or empties the computed properties so far. Default is True.
        '''
        
        #--------------------------------------------------------------------------------------------------------------------#
        #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        # Asserting that the input pressure parameter matches in type and length with the input parameter.
        if not isinstance(inpres, type(None)):
            assert(type(phi_q_vs) == type(inpres)), "Please provide an input pressure type that matches the input paramter type."
            if isinstance(inpres, np.ndarray) or isinstance(inpres, list):
                assert(len(inpres) == len(phi_q_vs)), "Please provide an input pressure list/array of the same length of the input parameter list/array."
        # Asserting that the input temperature parameter matches in type and length with the input parameter.
        if not isinstance(intemp, type(None)):
            assert(type(phi_q_vs) == type(intemp)), "Please provide an input temperature type that matches the input paramter type."
            if isinstance(intemp, np.ndarray) or isinstance(intemp, list):
                assert(len(intemp) == len(phi_q_vs)), "Please provide an input temperature list/array of the same length of the input parameter list/array."
                
        #--------------------------------------------------------------------------------------------------------------------#
        #--------------------------------------------------------BODY--------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
        
        #####
        ##### RECASTING TYPES
        #####
        # Recasting float input types to list types.
        if isinstance(phi_q_vs, float):
            phi_q_vs = [phi_q_vs]
            if inpres != None:
                inpres = [inpres]
            if intemp != None:
                intemp = [intemp]
        
        # Emptying output property lists.
        if reset:
            self.reset_props()        
        
        # Getting the list of turbine stages in cardinal order.
        stageproplist = sorted([_ for _ in dir(self) if "turbstage" in _], key=lambda x: int(x.split("turbstage")[-1]))
        
        # Declaring energetic-output-related lists of dimensional variables.
        torque = list()
        power = list()
        pst = list()
        ptt = list()
        
        # Declaring energetic-output-related lists of dimensionless variables.
        Phi = list()
        Pi = list()
        Psi = list()
        Eff = list()
        
        # Loop running over the set of input parameters.
        for phi_q_v in phi_q_vs:
            # Declaring auxiliar variables for computing the global energetic output values.
            inpower = 0
            outpower = 0
            torqueaux = 0
            pstaux = 0
            pttaux = 0
            Piaux = 0            
            Psiaux = 0            
            Effaux = 0
            # Declaring auxiliar lists for storing the dimensional energetic outputs of each turbine stage.
            torque_l = list()
            power_l = list()
            pst_l = list()
            ptt_l = list()
            # Declaring auxiliar lists for storing the dimensionless energetic outputs of each turbine stage.
            Pi_l = list()
            Psi_l = list()
            Eff_l = list()
            # Loop running over the turbine stages and applying the BEM method accordingly.
            for e, stageprop in enumerate(stageproplist):
                # Getting the current turbine stage.
                ts = getattr(self, stageprop)
                # Conditional for determining whether the current stage is the first stage.
                if e == 0:
                    # Setting the input pressures and temperatures.
                    if not isinstance(inpres, type(None)):
                        ts.fp.p = inpres[e]
                    if not isinstance(intemp, type(None)):
                        ts.fp.T = intemp[e]
                    # If it is, applying the BEM method with 'inputparam' set to the input parameter value.
                    ts.BEM(phi_q_v=phi_q_v, polars=polars, inputparam=inputparam, mode=mode)
                    # Appending the dimensionless flow-rate varaible to the 'Phi' list.
                    Phi.append(ts.it.Phi)
                else:
                    # If it is not, then getting the previous turbine stage instance.
                    tsprev = getattr(self, stageproplist[e - 1])
                    # Computing the input pressure to the current stage by substracting the total-to-total pressure losses to the previous stage's input pressure.
                    ts.fp.p = tsprev.fp.p - tsprev.eout.p_tot_to_tot
                    # Performing a deep copy of the previous stage's output velocity triangle, and setting it equal to the current stage's input triangle.
                    ts.it = copy.deepcopy(tsprev.ot)
                    # Calling the 'BEM' method for the current stage with 'inputparam' set to 'it'.
                    ts.BEM(phi_q_v=phi_q_v, polars=polars, inputparam='it', mode=mode)
                # Adding the dimensional energetic outputs to the auxiliar variables.
                # Adding the dimensionless energetic outputs to the auxiliar variables.
                Piaux += ts.eout.Pi                
                Psiaux += ts.eout.Psi                
                inpower += ts.eout.pvisc + ts.eout.pk + ts.eout.power
                outpower += ts.eout.power
                torqueaux += ts.eout.tau
                pstaux += ts.eout.p_st_to_st
                pttaux += ts.eout.p_tot_to_tot
                # Appending the dimensional energetic output of the turbine stage to the corresponding list.
                torque_l.append(ts.eout.tau)
                power_l.append(ts.eout.power)
                pst_l.append(ts.eout.p_st_to_st)
                ptt_l.append(ts.eout.p_tot_to_tot)
                # Appending the dimensionless energetic outputs of the turbine stage to the corresponding list.
                Pi_l.append(ts.eout.Pi)
                Psi_l.append(ts.eout.Psi)
                # Conditional for setting the efficiency value to 0 in case the stage's value falls outside the range [0, 1].
                if 0 <= ts.eout.eff <= 1:
                    Eff_l.append(ts.eout.eff)
                else:
                    Eff_l.append(0)
            # Appending the computed global dimensional energetic outputs to the corresponding lists.
            torque.append(torqueaux)
            power.append(outpower)
            pst.append(pstaux)
            ptt.append(pttaux)
            # Appending the computed global dimensionless energetic outputs to the corresponding lists.
            Pi.append(Piaux)
            Psi.append(Psiaux)
            # Conditional for setting the efficiency value to 0 in case the computed value falls outside the range [0, 1].
            if 0 <= outpower/inpower <= 1:
                Eff.append(outpower/inpower)
            else:
                Eff.append(0)
            # Adding the list of each turbine stage's dimensional energetic outputs to the corresponding output variable.
            self.torque_stages.append(torque_l)
            self.power_stages.append(power_l)
            self.pst_stages.append(pst_l)
            self.ptt_stages.append(ptt_l)
            # Adding the list of each turbine stage's dimensionless energetic outputs to the corresponding output variable.
            self.Pi_stages.append(Pi_l)
            self.Psi_stages.append(Psi_l)
            self.Eff_stages.append(Eff_l)
        
        if reset:
            # Transposing list of lists.        
            self.torque_stages = [list(i) for i in zip(*self.torque_stages)]
            self.power_stages = [list(i) for i in zip(*self.power_stages)]
            self.pst_stages = [list(i) for i in zip(*self.pst_stages)]
            self.ptt_stages = [list(i) for i in zip(*self.ptt_stages)]
            self.Pi_stages = [list(i) for i in zip(*self.Pi_stages)]
            self.Psi_stages = [list(i) for i in zip(*self.Psi_stages)]
            self.Eff_stages = [list(i) for i in zip(*self.Eff_stages)]
            # Setting the global dimensional energetic output properties of the turbine to the stored lists. 
            self.torque = torque
            self.power = power
            self.pst = pst
            self.ptt = ptt
            # Setting the global dimensionless energetic output properties of the turbine to the stored lists.
            self.Phi = Phi
            self.Pi = Pi
            self.Psi = Psi
            self.Eff = Eff
        else:
            # Transposing list of lists.        
            self.torque_stages += [list(i) for i in zip(*self.torque_stages)]
            self.power_stages += [list(i) for i in zip(*self.power_stages)]
            self.pst_stages += [list(i) for i in zip(*self.pst_stages)]
            self.ptt_stages += [list(i) for i in zip(*self.ptt_stages)]
            self.Pi_stages += [list(i) for i in zip(*self.Pi_stages)]
            self.Psi_stages += [list(i) for i in zip(*self.Psi_stages)]
            self.Eff_stages += [list(i) for i in zip(*self.Eff_stages)]
            # Setting the global dimensional energetic output properties of the turbine to the stored lists. 
            self.torque += torque
            self.power += power
            self.pst += pst
            self.ptt += ptt
            # Setting the global dimensionless energetic output properties of the turbine to the stored lists.
            self.Phi += Phi
            self.Pi += Pi
            self.Psi += Psi
            self.Eff += Eff
            
##########################################################################################################################################
##########################################################################################################################################
#####################################################SPECIFIC FUNCTIONS/METHODS###########################################################
##########################################################################################################################################
##########################################################################################################################################

#--------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------polar_database_load() METHOD--------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def polar_database_load(filename: str='POLARS/POLARS',
                        filepath: str=os.getcwd(),
                        pick: bool=False) -> dict:
    '''Reads a given polar database from an Excel file, or loads a previously saved
    and serialized Pickle file.
    
    The Excel file containing the polars should be arranged in a particular manner.
    Specifically, two Excel sheets should be devoted to each of the considered airfoils: one
    for the lift coefficients, and one for the drag coefficients. Thus, assuming that a workbook
    contains merely the polars corresponding to a NACA0015 airfoil, the workbook should contain
    two worksheets named, respectively, 'NACA0015CL' and 'NACA0015CD'. Those worksheets should
    contain, each, a first column with the angle-of-attack, and a set of columns with lift and
    drag coefficients obtained for whatever the Reynolds numbers employed.
    Thus, for example:
    
    ------------------WORKBOOK "POLARSNACA0015.xlsx"------------------
    ----------------------WORKSHEET "NACA0015CL"----------------------
    | COL1 | COL2 | COL3 | COL4 | ............................| COLN |
    |   0  | Re1  |  Re2 |  Re3 | ............................|  Rem |
    ------------------------------------------------------------------
    |  a1  | cl11 | cl21 | cl31 | ............................| clm1 |
    |  a2  | cl12 | cl22 | cl32 | ............................| clm2 |
    |  a3  | cl13 | cl23 | cl33 | ............................| clm3 |
                                ..........                            
                                ..........                            
                                ..........                            
                                ..........                            
                                ..........                            
    |  az  | cl1z | cl2z | cl3z | ............................| clmz | 
    ----------------------WORKSHEET "NACA0015CD"----------------------
    | COL1 | COL2 | COL3 | COL4 | ............................| COLN |
    |   0  | Re1  |  Re2 |  Re3 | ............................|  Rem |
    ------------------------------------------------------------------
    |  a1  | cd11 | cd21 | cd31 | ............................| cdm1 |
    |  a2  | cd12 | cd22 | cd32 | ............................| cdm2 |
    |  a3  | cd13 | cd23 | cd33 | ............................| cdm3 |
                                ..........                            
                                ..........                            
                                ..........                            
                                ..........                            
                                ..........                            
    |  az  | cd1z | cd2z | cd3z | ............................| cdmz |
    
    As observed, the worksheets begin with a '0' in the cel A11, and the rest of the first
    row corresponds to the Reynolds numbers at which the coefficients are obtained.
    The data is stored in columns, each column corresponding to the coefficients that
    correspond to the angle-of-attack stored in the first column.
    For an illustrative example, see 'POLARS.xlsx'.    
    
    **parameters**, **return**, **return types**:

        :param filename: name of the file (Excel/Pickle) to be read. Default is 'POLARS'
        :param filepath: path on which the file is located. Default is current directory.
        :param pickle: boolean flag indicating whether the data is to be loaded from a Pickle
        file. Default is False, meaning that an Excel file is read.        
        :return: dictionary containing read polar data.
        :rtype: dict.
    '''

    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Instantiating Excel and Pickle file names with corresponding extensions.
    file_path_name_xl = "\\".join([filepath, filename + ".xlsx"])
    file_path_name_p = "\\".join([filepath, filename + ".p"]) if filepath != "" else filepath + filename + ".p"
    ## Conditional for determining which type file is read.
    # If 'pick' is True, then the corresponding Pickle file is loaded (previously read from an Excel file).
    if pick:
        with (open(file_path_name_p, "rb")) as pickle_file:
            # Return statement.
            return pickle.load(pickle_file)
    # If 'pick' is False, then the data is loaded from an Excel file.
    else:
        # Instantiating 'wb_obj' as an Excel workbook object.
        wb_obj = openpyxl.load_workbook(file_path_name_xl)
        # Instantiating 'data' as the data dictionary on which to dump the read datasets.
        data = {}
        # Loop running over the sheets of the Excel workbook.
        for sheetname in wb_obj.sheetnames:
            # The 'sheet' variable addresses the sheet corresponding to the current iteration.
            sheet = wb_obj[sheetname]
            # The 'col_names' is a variable used for assigning names to the datasets that would be
            # read. As the datasets would be either angles-of-attack, and Reynolds-dependent lift
            # and drag coefficients, the 'col_names' variable stores either a string 'angle' (meaning
            # that the corresponding dataset refers to angles-of-attack) or the Reynolds number at
            # which the stored coefficients have been obtained.
            col_names = list()
            # Loop running over the first row to get the names of the columns.
            for column in sheet.iter_cols(1, sheet.max_column):
                col_names.append(column[0].value if column[0].value != 0 else 'angle')
            # Creating the 'datasheet' dictionary that will store the datasets.
            datasheet = {}
            # Loop running over the columns (beginning from the second row), and storing the values
            # on the corresponding datasheet column.
            for _, col in enumerate(sheet.iter_cols(min_row=2, values_only=True)):
                datasheet[col_names[_]] = list(col)
            ## Extending/reversing the datasets and reordering them for having complete in a
            ## negative-to-positive angle-of-attack arrangement. The conditional intends to distinguish
            ## between the cases considering the provided angle-of-attack range.
            # If the provided angle-of-attack range is below 180º, then the extension/reversing consists
            # of reversing the angles and lift coefficients, and mirroring the drag coefficients
            # symmetrically.
            if datasheet['angle'][-1] <= 180:
                # The loop runs over the datasheet keys.
                for key in datasheet.keys():
                    # The angular case is treated separately, as it always consist of a reversing operation.
                    if 'angle' == key:
                        datasheet[key] = [-_ for _ in datasheet[key][:0:-1]] + datasheet[key]
                    else:
                        # The lift case consists of a reversing operation.
                        if "CD" not in sheetname:
                            datasheet[key] = [-_ for _ in datasheet[key][:0:-1]] + datasheet[key]
                        # The drag case consists of a mirroring operation.
                        else:
                            datasheet[key] = datasheet[key][:0:-1] + datasheet[key]
                    # Casting the result into array format for easing further operations.
                    datasheet[key] = np.array(datasheet[key])
            # If the provided angle-of-attack range is above 180º, then the extension/reversing consist
            # of substracting the constant 360 to the angular values, for which a slicing of the
            # list is necessary.
            elif datasheet['angle'][-1] > 180:
                # Getting the index at which the angular value surpasses 180º.
                index = [_ >= 180 for _ in datasheet['angle']].index(True)
                # The loop runs over the datasheet keys.
                for key in datasheet.keys():
                    # The angular cas is treated separately, as it always consist of a reversing operation.
                    if 'angle' == key:
                        # The original dataset is splitted into positive and (about-to-be-)negative parts.
                        pos_part = datasheet[key][:index]
                        neg_part = datasheet[key][index:]
                        # The result dataset consists of substracting 360 to each value of the negative part,
                        # whereas the positive part remains.
                        datasheet[key] = [_ - 360 for _ in neg_part] + pos_part
                    # In the rest of the cases, no operations are needed. The negative part is just set
                    # in front of the positive part, to comply with the ascending ordering of the
                    # angle-of-attack.
                    else:
                        datasheet[key] = datasheet[key][index:] + datasheet[key][:index]
                    # Casting the result into array format for easing further operations.
                    datasheet[key] = np.array(datasheet[key])
                                                    
            # Setting the corresponding entrance on the 'data' variable to the datasheet.
            data[sheetname] = datasheet            
            
        #--------------------------------------------------------------------------------------------------------------------#
        #-------------------------------------------------------RETURN-------------------------------------------------------#             
        #--------------------------------------------------------------------------------------------------------------------#
            
        # Storing the file as a pickle, and returning the data-dictionary.
        with open(file_path_name_p, 'wb') as pickle_file:
            pickle.dump(data, pickle_file)
            return data

#--------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------double_interp() METHOD-----------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def double_interp(x1: float,
                  y1: float,
                  x2: float,
                  y2: float,
                  x: float,
                  y: float,
                  f11:float,
                  f12: float,
                  f21: float,
                  f22: float) -> float:
    '''Double interpolation function given two input variables.

    **parameters**, **return**, **return types**::
    
        :param x1: minimum value of the first variable.
        :param y1: minimum value of the second variable.
        :param x2: maximum value of the first variable.
        :param y2: maximum value of the second variable.
        :f11: value of the function to interpolate at (x1, y1).
        :f12: value of the function to interpolate at (x1, y2).
        :f21: value of the function to interpolate at (x2, y1).
        :f22: value of the function to interpolate at (x2, y2).
        :return: interpolated value.
        :rtype: real (float).   
    '''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Checks whether a single or double interpolation is required. If x1=x2 or y1=y2, then
    # a single interpolation formula is employed.
    if x1 == x2:
        # Single interpolation formula in case x1=x2.
        interp = f11 + ((f12 - f11)/(y2 - y1))*(y - y1)
    elif y1 == y2:
        # Single interpolation formula in case y1=y2.
        interp = f11 + ((f21 - f11)/(x2 - x1))*(x - x1)
    else:
        # Terms entering the double interpolation formula.
        term1 = f11*(x2 - x)*(y2 - y)
        term2 = f21*(x - x1)*(y2 - y)
        term3 = f12*(x2 - x)*(y - y1)
        term4 = f22*(x - x1)*(y - y1)        
        # Double interpolation formula; checkable in
        # https://en.wikipedia.org/wiki/Bilinear_interpolation
        interp = (1/((x2 - x1)*(y2 - y1)))*(term1 + term2 + term3 + term4)
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Return statement.
    return interp

#--------------------------------------------------------------------------------------------------------------------#
#----------------------------------------------dynamic_fitting() METHOD----------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def dynamic_fitting(x: np.ndarray,
                    y: np.ndarray,
                    mindeg: int=4,
                    resmin: float=1e-3,
                    message: bool=True) -> np.ndarray:
    '''Dynamic polynomial fitting function.

    **parameters**, **return**, **return types**::
    
        :param x: x component of data to be fit.
        :param y: y component of data to be fit.
        :param mindeg: minimum degree of polynomial to be used as fitting. Default is 4.
        :param resmin: minimum residual to tolerate. Default is 1e-3.
        :param message: boolean flag for showing message if polynomial fitting fails. Default is True.
        :return: fitted array.
        :rtype: ndarray.   
    '''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # While loop that runs until the minimum tolerance is reached or fitting fails.
    while True:
        # Calling polyfit with a specific polynomial degree (it is minimum on the first step),
        # and 'full' mode for getting the residuals.
        coefsres = np.polyfit(x, y, mindeg, full=True)
        # Setting fitting coefficients to 'coefs' variable.
        coefs = coefsres[0]
        # Setting residuals of fitting to 'res' variable.
        res = coefsres[1]
        # Conditional for checking if the polynomial fitting provides a valid set of residuals.
        if len(res) == 1:
            # Conditional for checking if the tolerance is reached.
            if res < resmin:
                # Break if it is reached.
                break
            # Conditional for checking if the maximum allowable polynomial degree is reached.
            elif mindeg >= len(x)//2:
                # Augment polynomial degree if it is not.
                mindeg += 1
                # Continue to next iteration step.
                continue
            # Break if maximum allowable polynomial degree is reached.
            elif mindeg >= len(x)//2:
                break
            # Break statement for exiting while loop if any of the stopping conditions
            # (tolerance, maximum degree) is reached.
            break
        # If the provided residual length is not valid, exit loop.
        elif len(res) > 1:
            break
        # If the while loop is broken due to maximum polynomial degree overcoming, then send message.        
        elif mindeg >= len(x)//2:
            # Show message if the boolean flag is True.
            if message:
                print("maximum degree of polynomial reached (" + str(mindeg) + ") without achieving tolerance.")
            # Break statement for exiting the loop.
            break
        else:
            # If none of the breaking statement is reached, then augment polynomial degree.
            mindeg += 1
    # Setting 'yfit' variable to a null array of size len(x).
    yfit = np.zeros(len(x))
    # Performing fitting based on fitting coefficient value.s
    for e, _ in enumerate(coefs[::-1]):
        yfit += coefs[len(coefs) - 1 - e]*x**(e)
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Return statement.
    return yfit

#--------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------knot_catcher() METHOD------------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def knot_catcher(x: Union[list, np.ndarray], y: Union[list, np.ndarray], retfit: bool=False, refine=True) -> list:
    '''Function for catching the knots on an x-y function.
    
    **parameters**, **return**, **rtypes**::
    
    :param x: x-wise array or list of function.
    :param y: y-wise array or list of function.
    :param retfit: boolean indicating whether the refined, fitted curve is returned. Default is False.
    :param refine: boolean indicating whether a refinement is to be performed on the input data. Default is True.
    :return: list comprising: # of knots, x-coordinates of knots, y-coordintaes of knots,
    fitted x-coordinates, fitted y-coordinates.
    :rtype: list(int, list, list, np.ndarray, np.ndarray)
    '''

    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#         
    
    #####
    ##### DETECTING THE REVERSALS ON X- AND Y-WISE COORDINATES
    #####
    # Difference array on x-wise coordinates.
    diffx = np.diff(x)
    # Boolean array on the differences of x-wise coordinates; 0 if the difference is positive; 1 otherwise.
    diffboolx = [0] + [0 if _ > 0 else 1 for _ in diffx]
    # The 'revxcount' counts the number of times the x-wise coordinate reverses, by computing the difference of the 'diffboolx' list.
    revxcount = [_ != 0 for _ in np.diff(diffboolx)].count(True)
    # Difference array on y-wise coordinates.
    diffy = np.diff(y)
    # Boolean array on the differences of y-wise coordinates; 0 if the difference is positive; 1 otherwise.
    diffbooly = [0] + [0 if _ > 0 else 1 for _ in diffy]
    # The 'revycount' counts the number of times the y-wise coordinate reverses, by computing the difference of the 'diffbooly' list.
    revycount = [_ != 0 for _ in np.diff(diffbooly)].count(True)
    # Insantiating the 'inds' variable.
    inds = list()
    # If the number of x- and y-wise reversals is greater than 1 (both of them), then the existence of a knot is possible (or knots).    
    if revxcount and revycount:
        # Getting the indices on which the x- and y-wise coordinates reverse.
        revxinds = [e for e, _ in enumerate(np.diff(diffboolx)) if _ != 0]
        revyinds = [e for e, _ in enumerate(np.diff(diffbooly)) if _ != 0]
        # Sorting the indices and storing them on the 'inds' variable.
        inds = sorted(revxinds + revyinds)
    else:
        return [0, [], [], np.array([]), np.array([])]
    # Prepending a '0' at the 'inds' variable; thus, its first index retrieves the first elements of the x- and y-wise coordinates.
    inds.insert(0, 0)
    # Appending a 'len(x)' at the 'inds' variable; thus, its last index retrieves the last elements of the x- and y-wise coordinates.
    inds.append(len(x))
    # Removing duplicated elements and sorting list.
    inds = sorted(list(set(inds)))

    #####
    ##### DECLARING VARIABLES FOR FUNCTION-REFINEMENT AND FUNTION-CHUNKING
    ####
    # If 'refine' input boolean is 'True', then it is necessary to perform a refinement on the input data.
    if refine:
        # The 'refpoints' addresses the refinement points by which the input function is to be refined.
        # The refinement considers '10*len(x)' points (tenfold refinement) if such a number is greater
        # than 1000; otherwise, the refinement is carried out with 1000 points.
        refpoints = 10*len(x) if 10*len(x) > 1000 else 1000
        # Refining the x-wise coordinate and storing it in the 'xrefined' variable.
        xrefined = np.linspace(min(x), max(x), refpoints)
    # Otherwise, the 'xrefined' variable is simply set to the input 'x' data.
    else:
        xrefined = x
    # The 'xchunks' chunks the x-wise coordinate into pieces for which there are neither x- nor y-reversals.
    xchunks = [x[inds[e]:inds[e + 1]] for e, _ in enumerate(inds[:-1])]
    # The 'ychunks' chunks the y-wise coordinate into pieces for which there are neither x- nor y-reversals.
    ychunks = [y[inds[e]:inds[e + 1]] for e, _ in enumerate(inds[:-1])]
    # The 'xfitchunks' variable addresses a list that will store the refined x-wise chunks.
    xfitchunks = list()
    # The 'yfitchunks' variable addresses a list that will store the refined y-wise chunks.
    yfitchunks = list()
    # The 'argss' variable addresses a list that will store the arguments (indices) for refining the chunks.
    argss = list()
    # The 'revflags' variable addresses a boolean list that tells whether the considered chunks are in reversed order.
    revflags = list()
    
    #####
    ##### FITTING THE CHUNKS
    #####
    # Loop running over the chunks for performing the refined fitting.
    for e, _ in enumerate(xchunks):
        if len(xchunks[e]) == 1:
            xfitchunks.append(xchunks[e])
            yfitchunks.append(ychunks[e])
        else:
            # Getting the indices at which the 'xrefined' variable gets closest to the input x-wise coordinate.        
            args = sorted((np.argmin(np.abs(np.array([xchunks[e][0] - _ for _ in xrefined]))), np.argmin(np.abs(np.array([xchunks[e][-1] - _ for _ in xrefined])))))
            # Appending those arguments to 'argss'.
            argss.append(args)
            # Appending the refined x-wise chunk portion to 'xfitchunks'.
            xfitchunks.append(list(xrefined[args[0]: args[1]]))
            # Checking whether the chunks are in reversed x-wise order.
            if xchunks[e][-1] < xchunks[e][0]:
                # If so, reverting the chunks.
                xchunks[e] = xchunks[e][::-1]
                ychunks[e] = ychunks[e][::-1]
                # Appending a 'True' value to the 'revflags' variable.
                revflags.append(True)
            else:
                # Appending a 'False' value to the 'revflags' variable.
                revflags.append(False)
            # Performing the y-wise interpolation on the chunk portion.
            f = interp1d(xchunks[e], ychunks[e], kind='slinear', fill_value='extrapolate')
            # Appending the refined y-wise chunk portion to 'yfitchunks'.
            yfitchunks.append(list(f(xfitchunks[-1])))
    
    #####
    ##### KNOT CATCHING
    #####
    # The 'knots' variable will store the number of knots in the function; its initial value is 0.
    knots = 0
    # The 'knotsx' and 'knotsy' lists will store the x- and y-coordinates of the knots.
    knotxs = list()
    knotys = list()
    # Loop ruunning over the chunks for catching the knots.
    for e, _ in enumerate(xfitchunks):        
        # If the first chunk is being considered or the length of the chunk is 1 (return point), then skip (in a sequential analysis, knots will show up after the second chunk portion).
        if e == 0 or len(xfitchunks[e]) <= 1:
            continue
        # Otherwise, analyse for intersections between current chunk portion and previous chunks.
        else:
            # Loop running over the previous chunks.
            for f, __ in enumerate(xfitchunks[:e]): 
                # If the length of the chunk is 1 (return point), then skip.
                if len(xfitchunks[f]) <= 1:
                       continue
                # The 'minmincond1' boolean checks whether the minimum value of the considered previous chunk is lower than the minimum value of the current chunk.                
                minmincond1 = min(xfitchunks[f]) <= min(xfitchunks[e])
                # The 'minmincond2' boolean checks whether the minimum value of the current chunk is lower than the minimum value of the considered previous chunk.
                minmincond2 = min(xfitchunks[e]) <= min(xfitchunks[f])
                # The 'maxmincond1' boolean checks whether the maximum value of the considered previous chunk is lower than the minimum value of the current chunk.
                maxmincond1 = max(xfitchunks[f]) <= min(xfitchunks[e])
                # The 'maxmincond2' boolean checks whether the maximum value of the current chunk is lower than the minimum value of the considered previous chunk.
                maxmincond2 = max(xfitchunks[e]) <= min(xfitchunks[f])                
                
                #####
                ##### GETTING THE BEGINNING OF THE OVERLAPPING PORTION BETWEEN CHUNKS
                #####
                # If 'maxmincond1' or 'maxmincond2' are True, it means that the chunks do not overlap. Skipping to the next previous chunk.
                if maxmincond1 or maxmincond2:
                    continue
                # If 'minmincond1' is True and 'maxmincond1' is False,  it means that the curves overlap, at least, partially.
                elif minmincond1 and not maxmincond1:
                    # Due to the discreteness of the input function, two overlapping x-coordinates will ensue: one for the considered previous chunk, and one for the current chunk.
                    # The x-wise coordinate at which the overlapping begins for the considered previous chunk is obtained.
                    minargf = [_ - xfitchunks[e][0] >= 0 for _ in xfitchunks[f]].index(True) - 1
                    # The x-wise coordinate at which the overlapping begins for the current chunk is obtained.
                    if xfitchunks[f][minargf] not in xfitchunks[e]:
                        continue
                    minarge = xfitchunks[e].index(xfitchunks[f][minargf])
                # If 'minmincond2' is True and 'maxmincond2' is False,  it means that the curves overlap, at least, partially.
                elif minmincond2 and not maxmincond2:
                    # Due to the discreteness of the input function, two overlapping x-coordinates will ensue: one for the considered previous chunk, and one for the current chunk.
                    # The x-wise coordinate at which the overlapping begins for the considered previous chunk is obtained.
                    minarge = [_ - xfitchunks[f][0] >= 0 for _ in xfitchunks[e]].index(True) - 1
                    # The x-wise coordinate at which the overlapping begins for the current chunk is obtained.
                    if xfitchunks[e][minarge] not in xfitchunks[f]:
                        continue
                    minargf = xfitchunks[f].index(xfitchunks[e][minarge])
                
                #####
                ##### GETTING THE END OF THE OVERLAPPING PORTION BETWEEN CHUNKS
                #####
                # If the last x-wise coordinate of the current chunk is lower than the last x-wise coordinate of the considered previous chunk.
                if xfitchunks[e][-1] < xfitchunks[f][-1]:
                    # The end of the overlapping section is determined by the current chunk.
                    maxarge = len(xfitchunks[e]) - 1
                    # The x-wise coordinate corresponding to the end of the overlapping section for the considered previous chunk is obtained from the value of the last x-wise coordinate of the current chunk.
                    maxargf = xfitchunks[f].index(xfitchunks[e][maxarge])
                # If the last x-wise coordinate of the considered previous chunk is lower than the last x-wise coordinate of the current chunk.                    
                else:
                    # The end of the overlapping section is determined by the considered previous chunk.
                    maxargf = len(xfitchunks[f]) - 1
                    # The x-wise coordinate corresponding to the end of the overlapping section for the current chunk is obtained from the value of the last x-wise coordinate of the considered previous chunk.
                    maxarge = xfitchunks[e].index(xfitchunks[f][maxargf])                                                
                # In case 'minarge' and 'maxarge' match, or 'minargf' and 'maxargf' coincide, then no overlap exists between chunks. Skipping iteration.
                if minarge == maxarge or minargf == maxargf:
                    continue                
                
                #####
                ##### GETTING THE Y-WISE OVERLAPPING SECTIONS AND EXISTING INTERSECTIONS
                #####
                # Getting the y-wise overlapping section for the current chunk.
                split_yfitchunkse = yfitchunks[e][minarge:maxarge]
                # Getting the y-wise overlapping section for the considered previous chunk.
                split_yfitchunksf = yfitchunks[f][minargf:maxargf]
                # Conditional for checking whether the first element of the y-wise coordinate of the current chunk is lower than the first element of the y-wise coordinate of the considered previous chunk.
                if split_yfitchunkse[0] < split_yfitchunksf[0]:
                    # If so, the passing to negative 'diff' values, 'diff' corresponding to the substraction of the y-wise values of the overlapping, means that an intersection occurs.
                    diff = np.array(split_yfitchunksf) - np.array(split_yfitchunkse)
                else:
                    # Otherwise, the passing to negative 'diff' values, 'diff' corresponding to the substraction of the y-wise values of the overlapping, means that an intersection occurs.
                    diff = np.array(split_yfitchunkse) - np.array(split_yfitchunksf)
                # The 'neginds' variable determines whether negative 'diff' values are present.
                neginds = np.argwhere(diff < 0)
                # If so:
                if len(neginds) > 0:
                    # The number of knots is increased by 1 (a knot exist between the considered function chunks).
                    knots += 1
                    # The index at which the intersection occurs is retrieved from 'neginds'.
                    knotind = neginds[0][0]
                    # The x-wise coordinate of the knot for the current chunk is obtained.
                    knotxe = xfitchunks[e][knotind]
                    # The y-wise coordinate of the knot for the considered previous chunk is obtained.
                    knotye = split_yfitchunkse[knotind]
                    # The 'indf' is the index at which the intersection occurs for the considered previous chunk.
                    indf = xfitchunks[f].index(knotxe)
                    # The x-wise coordinate of the knot for the considered previous chunk is obtained.
                    knotxf = xfitchunks[f][indf]
                    # The y-wise coordinate of the knot for the considered previous chunk is obtained.
                    knotyf = yfitchunks[f][indf]
                    # The obtained x-wise knot coordaintes are appended to the 'knotxs' list.
                    knotxs.append([knotxe, knotxf])
                    # The obtained y-wise knot coordaintes are appended to the 'knotys' list.
                    knotys.append([knotye, knotyf])
    # The 'retlist' variable stores the return data list; it is set to 'knots', 'knotxs' and 'knotys'.
    retlist = list([knots, knotxs, knotys])
    # If 'retfit' is true, it is necessary to concatenate the fitted chunks.
    if retfit:
        # Reversing fitted chunks if 'revflag' is True.
        for e, revflag in enumerate(revflags):
            if revflag:
                xfitchunks[e] = xfitchunks[e][::-1]
                yfitchunks[e] = yfitchunks[e][::-1]
        # Concatenating fitted chunks.
        retlist.append(sum(xfitchunks, []))
        retlist.append(sum(yfitchunks, []))
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#     
    
    # Return statement.
    return retlist

#--------------------------------------------------------------------------------------------------------------------#
#-----------------------------------------------knot_trimmer() METHOD------------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def knot_trimmer(x: Union[list, np.ndarray], y: Union[list, np.ndarray], retdiscr: bool=False) -> list:
    '''Function for catching the knots on an x-y function.
    
    **parameters**, **return**, **rtypes**::
    
    :param x: x-wise array or list of function.
    :param y: y-wise array or list of function.
    :param retdiscr: boolean indicating whether the discrete trimmed curve is returned. Default is False.
    :return: (fitted x-coordinates, fitted y-coordinates), (discrete x-coordinates, discrete y-coordinates),
    (discrete function trim initial trim index, discrete function final trim index).
    :rtype: list(list, list, list)
    '''    

    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#     
    
    #####
    ##### DEEP-COPYING DISCRETE INPUT X- AND Y-COORDINATES
    #####
    # Deepcopying x- and y-coordinates.
    discr_trim_xs = copy.deepcopy(x)
    discr_trim_ys = copy.deepcopy(y)
    # Instantiating list for storing indices at which the discrete function is trimmed.
    discr_inds = list()
    
    #####
    ##### PROCESSING INPUT FUNCTION FOR KNOT DETECTION
    #####
    # Calling 'knot_catcher()' function with 'retfit' set to 'True' for getting the knot parameters of the input data.
    knot_params = knot_catcher(x=x, y=y, retfit=True)    
    # The number of knots of the current function is retrieved from 'knot_params'.
    knots = knot_params[0]
    # Setting the x- and y-coordinates of the knots to 'knotxs' and 'knotys'.
    knotxs = knot_params[1]
    knotys = knot_params[2]
    # Setting the fitted x- and y-coordinates to 'fit_xs' and 'fit_ys'.
    fit_xs = knot_params[3]
    fit_ys = knot_params[4]
        
    #####
    ##### TRIMMING PROCESS
    #####
    # If there are knots in the function, then proceed with the trimming.
    if knots > 0:        
        # Loop for performing trimming operations while the number of knots on the function remains above 1.
        while knots >= 1:
            # Computing the average x- and y-coordinate of the first knot if necessary.
            knotxave = np.average(knotxs[0])
            knotyave = np.average(knotys[0])
            #####
            ##### LEFT AND RIGHT HAND-SIDE KNOT COORDINATE COLLAPSING
            #####
            # Due to the discreteness of the input function, the left and right hand-side coordinates of the nodes may vary.
            # Getting the left hand-side x-coordinates of the knot.
            indx0 = [_[0] for _ in np.argwhere(np.array([np.abs(_ - knotxs[0][0]) for _ in fit_xs]) == 0)]
            # Getting the right hand-side x-coordinates of the knot.
            indx1 = [_[0] for _ in np.argwhere(np.array([np.abs(_ - knotxs[0][1]) for _ in fit_xs]) == 0)]
            # Getting the left hand-side y-coordinates of the knot.
            indy0 = [_[0] for _ in np.argwhere(np.array([np.abs(_ - knotys[0][0]) for _ in fit_ys]) == 0)]
            # Getting the right hand-side y-coordinates of the knot.
            indy1 = [_[0] for _ in np.argwhere(np.array([np.abs(_ - knotys[0][1]) for _ in fit_ys]) == 0)]                        
            # Conditional for filtering on left hand-side x- and y-coordinates of the knot and getting a unique value.
            if any([_ in indx0 for _ in indy0]) or any([_ + 1 in indx0 for _ in indy0]) or any([_ - 1 in indx0 for _ in indy0]):
                if any([_ in indx0 for _ in indy0]):
                    ind0 = indy0[[_ in indx0 for _ in indy0].index(True)]
                elif any([_ + 1 in indx0 for _ in indy0]):
                    ind0 = indy0[[_ + 1 in indx0 for _ in indy0].index(True)]
                else:
                    ind0 = indy0[[_ - 1 in indx0 for _ in indy0].index(True)]
            # Conditional for filtering on the right hand-side x- and y-coordinates of the knot and getting a unique value.
            if any([_ in indx1 for _ in indy1]) or any([_ + 1 in indx1 for _ in indy1]) or any([_ - 1 in indx1 for _ in indy1]):
                if any([_ in indx1 for _ in indy1]):
                    ind1 = indy1[[_ in indx1 for _ in indy1].index(True)]
                elif any([_ + 1 in indx1 for _ in indy1]):
                    ind1 = indy1[[_ + 1 in indx1 for _ in indy1].index(True)]
                else:
                    ind1 = indy1[[_ - 1 in indx1 for _ in indy1].index(True)]
            # Sorting the unique left and right hand-side knot coordinates.
            inds = sorted([ind0, ind1])            
            #####
            ##### TRIMMING FITTED FUNCTION
            #####
            # Trimming the input functions.
            trim_xs = fit_xs[:inds[0]] + fit_xs[inds[1]:]
            trim_ys = fit_ys[:inds[0]] + fit_ys[inds[1]:]
            #####
            ##### TRIMMING DISCRETE FUNCTION
            #####
            # Getting signs of differences of initial point with respect to knot coordinates.
            signx1 = np.sign(knotxave - x[0])
            signy1 = np.sign(knotyave - y[0])
            # Getting signs of differences of final point with respect to knot coordinates.
            signx2 = np.sign(x[-1] - knotxave)
            signy2 = np.sign(y[-1] - knotyave)
            # Getting signs of differences of initial point with respect to knot coordinates.
            discr_indx0 = max([signx1*(_ - knotxave) > 0 for _ in x].index(True) - 1, [signy1*(_ - knotyave) > 0 for _ in y].index(True) - 1)
            discr_indx1 = min(len(x) - [signx2*(knotxave - _) > 0 for _ in x[::-1]].index(True) + 1, len(y) - [signy1*(knotyave - _) > 0 for _ in y[::-1]].index(True) + 1)
            # Appending indices to 'discr_inds' list.
            discr_inds.append(discr_indx0)
            discr_inds.append(discr_indx1)
            # Trimming discrete function and adding knot point.
            discr_trim_xs = discr_trim_xs[:discr_indx0] + [knotxave] + discr_trim_xs[discr_indx1:]
            discr_trim_ys = discr_trim_ys[:discr_indx0] + [knotyave] + discr_trim_ys[discr_indx1:]            
            #####
            ##### REPROCESSING FOR KNOT DETECTION
            #####
            # Calling 'knot_catcher()' function with 'retfit' set to 'True' and 'refine' set to 'False'.
            # It intends to check whether the trimming operation has removed the other knots as well.
            knot_params = knot_catcher(x=trim_xs, y=trim_ys, retfit=True, refine=False)
            # Updating the knot parameters.
            knots = knot_params[0]
            knotxs = knot_params[1]
            knotys = knot_params[2]
            fit_xs = knot_params[3]
            fit_ys = knot_params[4]
            # If the number of knots lies below 1, breaking.
            if knots <= 1:
                break
    # Otherwise, just setting the trimmed function to the input function.
    else:
        trim_xs = x
        trim_ys = y
    discr_inds.insert(0, 0)
    discr_inds.append(len(x))
    discr_inds = [[discr_inds[e], discr_inds[e + 1]] for e in range(0, len(discr_inds) - 1, 2)]
    # The 'retlist' variable stores the return data list; it is set to '[trim_xs, trim_ys]'.
    retlist = [[trim_xs, trim_ys]]
    # If 'retdiscr' is 'True', then the trimmed discrete input function is also returned, as well as the indices at which it is trimmed.
    if retdiscr:
        retlist.append([discr_trim_xs, discr_trim_ys])
        retlist.append(discr_inds)
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Return statement.
    return retlist

#--------------------------------------------------------------------------------------------------------------------#
#-------------------------------------------psi_phi_slope_fitter() METHOD--------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def psi_phi_slope_fitter(Phi: Union[list, np.ndarray],
                         Psi: Union[list, np.ndarray],
                         interpfit = False) -> tuple:
    '''Linear range slope calculation of the psi-phi functional relation of a Wells turbine.
    
     **parameters**, **returns**, **rtypes** ::
    
    :param Phi: dimensionless flow-parameter variable.
    :param Psi: dimensionless pressure-loss variable.
    :param interfit: boolean determining whether an interpolation-based fitting is to be performed on the input data. Default is False.
    :returns: average slope and ordinate origin of the fitted least-squared equation.
    :rtypes: tuple(float, float)
    '''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------# 
    
    # Asserting that the input variables have the same length.
    assert all([len(_) == len(Phi) for _ in [Phi, Psi]]), "Please provide input data of same length."
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#         
    
    #####
    ##### INTERPOLATION-BASED FITTING, IF NECESSARY
    #####
    if interpfit:        
        f = interp1d(Psi, Phi, kind='slinear', fill_value='extrapolate')
        Psi_fit = np.linspace(Psi[0], Psi[-1], 1000)
        Phi_fit = f(Psi_fit)
        Psi = Psi_fit
        Phi = Phi_fit
        
    #####
    ##### LEAST-SQUARE-BASED SLOPE FITTING ON THE PSI-PHI FUNCTIONAL RELATION
    #####
    # Computing discrete Psi-step.
    dPsi = (Psi[-1] - Psi[0])/len(Psi)
    # Instantiating a 'slopes' list that will store the different slope values.
    slopes = list()
    # Loop running over the 'Psifit' variable.
    for e, i in enumerate(Psi[:-3]):
        # Computing the least-square-based slope of the fitted Psi-Phi functional relation with a number of progressively less points, and adding such slope to 'slopes'.
        slopes.append(mt.least_squares(np.array(Psi[:-(e+1)]), np.array(Phi[:-(e+1)]))[0])
    # Reverting the 'slopes' list.
    slopes = slopes[::-1]
    # Getting the threshold value at which the dPsi-scaled slopes overcome the unit value, which means that the constancy condition is violated. Such a threshold marks the point at which the fitted Psi-Phi functional relation ceases to be linear.
    threshold_array = [_ > 1 for _ in np.abs(np.diff(slopes))/dPsi]
    if True in threshold_array:
        threshold = len(Psi) - [_ > 1 for _ in np.abs(np.diff(slopes))/dPsi].index(True)
    else:
        threshold = len(Psi) - 1
    # Instantiating 's0s' and 'y0s' variables that will store the slopes and origin-ordinates upon the threshold range.
    s0s = list()
    y0s = list()
    # Loop running over the reversed threshold range.
    for e in range(threshold, 3, -1):
        # Computing the least-square-based slope and origin-ordinate with a number of progressively lower points.
        s0, y0 = mt.least_squares(np.array(Psi[:e]), np.array(Phi[:e]))[:2]
        # Appending the computed slopes and origin-ordinates to the corresponding lists.
        s0s.append(s0)
        y0s.append(y0)
    # Computing the average slopes and origin-ordinates.
    s0ave = np.average(s0s)
    y0ave = np.average(y0s)
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#     
    
    # Return statement.
    return s0ave, y0ave

#--------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------stochastic_analysis() METHOD--------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def stochastic_analysis(Phi: Union[list, np.ndarray],
                        Psi: Union[list, np.ndarray],
                        Pi: Union[list, np.ndarray],
                        maxdev: float=0.2, n: int=30, try_trim=True) -> tuple:
    '''Gaussian-based stochastic analysis upon dimensionless turbine variables.
    
    **parameters**, **returns**, **rtypes** ::
    
    :param Phi: dimensionless flow-parameter variable.
    :param Psi: dimensionless pressure-loss variable.
    :param Pi: dimensionless power variable.
    :param maxdev: float number representing the maximum, normalized standard deviation by which
    to perform the stochastic analysis. Default is 0.2.
    :param n: integer representing the number of normalized standard deviations to employ for
    performing the stochastic analysis. Default is 30.
    :param try_trim: boolean for stating whether to try trimming the Psi-Pi functional. Default is True.
    :returns: a tuple containing four arrays: the array of maximum deviations, and the arrays of
    stochastic dimensionless flow-rates, powers and efficiencies, respectively.
    :rtypes: tuple(ndarray, ndarray, ndarray, ndarray)
    '''

    #--------------------------------------------------------------------------------------------------------------------#
    #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------# 
    
    # Asserting that the input variables have the same length.
    assert all([len(_) == len(Phi) for _ in [Psi, Pi]]), "Please provide input data of same length."

    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#             
    
    #####
    ##### KNOT-TRIMMING PSI-PI FUNCTIONAL RELATION
    #####    
    # The trimming is performed only if 'try_trim' is 'True'.
    if try_trim:
        # Calling knot_trimmer() function upon Psi-Pi functional relation for trimming any possible knot.
        trim_psipi, trim_discr_psipi, discr_inds = knot_trimmer(Psi, Pi, retdiscr=True) 
    else:
        # Setting 'trim_psipi' and 'trim_discr_psipi' to tuple corresponding to input (Psi, Pi) variables.
        trim_psipi = (Psi, Pi)
        trim_discr_psipi = (Psi, Pi)
    
    #####
    ##### INSTANTIATING DISCRETE AND FITTED (CONTINUOUS) DIMENSIONLESS VARIABLES
    #####
    # Instantiating discrete Psi variable.
    Psidiscr = np.array(trim_discr_psipi[0])
    # Instantiating fitted (continuous) Psi variable.
    Psifit = np.array(trim_psipi[0])
    # Computing discrete Psi-step.
    dPsi = (Psifit[-1] - Psifit[0])/len(Psifit)
    # Instantiating discrete Pi variable.
    Pidiscr = np.array(trim_discr_psipi[1])
    # Instantiating fitted (continuous) Pi variable.
    Pifit = np.array(trim_psipi[1])
    # Instantiating an empty list for storing the discrete Phi variable.
    Phidiscr = list()
    # Conditional for checking whether a trimming operation is necessary on the Phi variable.
    if try_trim:
        if len(discr_inds[0]) > 0:
            # Loop for trimming the Phi variable according to the already trimmed Pi variable.
            for e, discr_ind in enumerate(discr_inds):
                Phidiscr += Phi[discr_ind[0]:discr_ind[1]]
                # In case the trimming operation does not correspond to the last chunk, adding the average coordinates at both sides of the chunk to the Phi variable.
                if e != len(discr_inds) - 1:
                    Phidiscr += [np.average([Phi[discr_inds[e][1]], Phi[discr_inds[e + 1][0]]])]
    else:
        Phidiscr = Phi
    # Interpolating the Psi-Phi functional for fitting the Phi variable.
    f = interp1d(Psidiscr, Phidiscr, kind='slinear', fill_value='extrapolate')
    # Instantiating the fitted (continuous) Phi variable.
    Phifit = f(Psifit)
    
    # Least-square fitting and computation of average slope and ordinate-origin of the phi-psi functional relation.
    s0ave, y0ave = psi_phi_slope_fitter(Phifit, Psifit)    

    #####
    ##### GAUSSIAN-FUNCTION-BASED EXTENT DETERMINATION
    #####
    # Instantiating the 'devs' variable, an evenly-spaced array of standard deviations.
    devs = np.linspace(1e-2, maxdev, n)
    # Computing the maximum gaussian value based on the maximum standard deviation and the last element of 'Psifit'.
    maxgauss = np.exp(-Psifit[-1]**2/(2*maxdev**2))
    # Instantiating the 'Psimingaussval' variable and setting its value to the last element of 'Psifit'.
    Psimingaussval = Psifit[-1]
    # Instantiating the 'mingaussval' variable and setting its value to the outcome of the Gaussian function with the maximum deviation value and 'Psimingaussval'.
    mingaussval = np.exp(-Psimingaussval**2/(2*maxdev**2))
    # Instantiating the 'iPsi' variable.
    iPsi = 0
    # While loop for extending the dimensionless variables until the outcome of the gaussian function for the largest maximum deviation lies below 1e-4.
    while mingaussval > 1e-4:
        # Incrementing 'Psimingaussval' by 'dPsi'.
        Psimingaussval += dPsi
        # Incrementing 'iPsi' by 1.
        iPsi += 1
        # Computing 'mingaussval'.
        mingaussval = np.exp(-Psimingaussval**2/(2*maxdev**2))

    #####
    ##### EXTENDING DIMENSIONLESS VARIABLES
    #####
    # Instantiating 'Psifit_stoch_add' as the portion of curve required for extending the 'Psifit' variable until getting to the Gaussian limit.
    Psifit_stoch_add = [_ for _ in np.linspace(Psifit[-1], Psimingaussval, iPsi)]
    # Instantiating 'Psifit_stoch' as an extent of the 'Psifit' variable.
    Psifit_stoch = np.array([_ for _ in Psifit] + Psifit_stoch_add)
    # Instantiating 'Phifit_stoch' as the linear extent based on 'Psifit_stoch'.
    Phifit_stoch = s0ave*Psifit_stoch + y0ave
    # Instantiating 'Pifit_stoch' by setting the values of the extent portion to the last value of the 'Pifit' variable.
    Pifit_stoch = np.array([_ for _ in Pifit] + [Pifit[-1] for _ in Psifit_stoch_add])
    # Computing the available dimensionless power as the product of the extended and fitted Psi and Phi variables, and instantiating the 'PiAvailfit_stoch' variable.
    PiAvailfit_stoch = Psifit_stoch*Phifit_stoch

    #####
    ##### COMPUTING THE STOCHASTIC AVERAGES
    #####
    # Instantiating the 'avePhi' variable for storing the average value of the stochastic Phi variable.
    avePhi = list()
    # Instantiating the 'avePi' variable for storing the average value of the stochastic Pi variable.
    avePi = list()
    # Instantiating the 'aveEff' variable for storing the average value of the stochastic Eff variable.
    aveEff = list()
    # Loop running over the standard deviations.
    for dev in devs:
        # Computing the average value of the stochastic Phi variable.
        avePhi.append((1/(np.sqrt(2*np.pi)*dev))*simpson(np.exp(-Psifit_stoch**2/(2*dev**2))*Phifit_stoch, dx=dPsi))
        # Computing the average value of the stochastic Pi variable.
        avePi.append((1/(np.sqrt(2*np.pi)*dev))*simpson(np.exp(-Psifit_stoch**2/(2*dev**2))*Pifit_stoch, dx=dPsi))
        # Computing the average value of the stochastic PiAvail variable.
        avePiAvail = (1/(np.sqrt(2*np.pi)*dev))*simpson(np.exp(-Psifit_stoch**2/(2*dev**2))*PiAvailfit_stoch, dx=dPsi)
        # Computing the average value of the stochastic Eff variable.
        aveeff = avePi[-1]/avePiAvail
        if 0 < aveeff < 1:
            aveEff.append(avePi[-1]/avePiAvail)
        else:
            aveEff.append(0)

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#                
            
    # Return statement.
    return devs, avePhi, avePi, aveEff


#--------------------------------------------------------------------------------------------------------------------#
#---------------------------------------------pitch_optimization() METHOD--------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def pitch_optimization(turb,
                       polars: dict,
                       vxlist: list=list(range(1, 46)), angpitch: tuple=(-15, 0), angpitch_mode: str='decrease') -> tuple:
    '''Power-based pitch optimization upon a turbine, for a range of given input velocities and target pitch angles.
    
    **parameters**, **returns**, **rtypes** ::
    
    :param turb: an instantiated object of the 'turbine' class.
    :param polars: dimensionless pressure-loss variable.
    :param vxlist: input velocities at which to power-optimize the pitch. Default is a list of values ranging from 1 m/s to 30 m/s.
    :param angpitch: set of pitch angles to power-optimize. Default is a list of values ranging from -10º to 0º.
    :returns: a tuple containing five lists: the computed dimensionless flow-coefficients, optimum powers, optimum angular pitches,
    optimum dimensionless pressure losses, optimum dimensionless powers and optimum efficiencies.
    :rtypes: tuple(list, list, list, list, list, list)    
    '''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Asserting that the 'angpitch_mode' variable is a valid parameter.
    assert angpitch_mode in ['decrease', 'increase'], "Provide a valid anpitch_mode, either 'decrease' or 'increase'."

    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Creating auxiliary power and dimensionless flow-coefficient, pressure-coefficient, power-coefficient and efficiency for end-of-the-pitching-loop storage.
    pows_tot = list()
    Phis_tot = list()
    Psis_tot = list()
    Pis_tot = list()
    Effs_tot = list()
    Pitches_tot = list()
    
    # Declaring 'angpitch_bound' variable, which stores the value of the first angular pitch configuration to test according to 'angpitch_mode'.
    angpitch_bound = angpitch[1] if angpitch_mode == 'decrease' else angpitch[0]
    # Declaring 'angpitch_stop_bound' variable, which stores the value of the last angular pitch configuration to be tested according to 'angpitch_mode'.
    angpitch_stop_bound = angpitch[1] if angpitch_mode == 'increase' else angpitch[0]
    # Declaring 'var_sign' parameter, which stores the variational sign of the angular pitch according to 'angpitch_mode'.
    var_sign = -1 if angpitch_mode == 'decrease' else 1
    # Declaring 'var_pitch' boolean flag for checking whether a pitch-variation-based analysis is to be undertaken within the BEM-based loop.
    var_pitch = True
    # BEM-based loop according to the provided 'vxlist' list.
    for i in vxlist:
        # Setting turbine's angular pitch to 'angpitch_bound'.
        turb.turbstage1.gp.angpitch = angpitch_bound
        # Applying BEM.
        turb.BEM([i], polars=polars, inputparam='vx', mode='3DAD', reset=True)
        # Storing BEM-derived data in auxiliary variables.
        Phis_tot.append(turb.Phi[0])
        Psis_tot.append(turb.Psi[0])
        Pis_tot.append(turb.Pi[0])
        Effs_tot.append(turb.Eff[0])
        pows_tot.append(turb.turbstage1.eout.power)
        # Setting 'pow_max' variable to obtained output power.
        pow_max = turb.turbstage1.eout.power
        
        #####
        ##### CONDITIONAL FOR PITCH-VARIATION-BASED ANALYSIS
        #####
        # If 'var_pitch' is 'True', then perform pitch-variation-based analysis.
        if var_pitch:
            # Set 'stop_loop' variable to 'False'.
            stop_loop = False
            # While loop running while 'stop_loop' is 'False'.
            while not stop_loop:
                # Varying angular pitch of turbine's stage according to 'var_sign'.
                turb.turbstage1.gp.angpitch += var_sign
                # Re-applying BEM with varied angular pitch.
                turb.BEM([i], polars=polars, inputparam='vx', mode='3DAD', reset=True)
                # Conditional for checking whether obtained output power is larger than 'pow_max'.
                if turb.turbstage1.eout.power < pow_max:
                    # If it is not, then no variation is required in the angular pitch yet, and pre-variational pitch angle is added to 'Pitches_tot' auxiliary variable, apart form stopping the loop.
                    Pitches_tot.append(turb.turbstage1.gp.angpitch - var_sign)
                    stop_loop = True
                else:
                    # Otherwise, recasting last values of auxiliary variables to output variables of current BEM application.
                    Phis_tot[-1] = turb.Phi[0]
                    Psis_tot[-1] = turb.Psi[0]
                    Pis_tot[-1] = turb.Pi[0]
                    Effs_tot[-1] = turb.Eff[0]
                    pows_tot[-1] = turb.turbstage1.eout.power
                    # Resetting 'pow_max' variable to current output power value.
                    pow_max = turb.turbstage1.eout.power 
                    # Resetting 'angpitch_bound' to current pitch angle.
                    angpitch_bound = turb.turbstage1.gp.angpitch
                    # Conditional for checking whether 'angpitch_stop_bound' has been reached. If so, then cancelling further pitch-variation-based analyses.
                    if (var_sign < 0 and angpitch_bound < angpitch_stop_bound) or (var_sign > 0 and angpitch_bound > angpitch_stop_bound):
                        var_pitch = False
        else:
            # If the pitch-variation-based analysis is cancelled out, then appending 'angpitch_stop_bound' to 'Pitches_tot'.
            Pitches_tot.append(angpitch_stop_bound)   

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Return statement.
    return Phis_tot, pows_tot, Pitches_tot, Psis_tot, Pis_tot, Effs_tot

##########################################################################################################################################
##########################################################################################################################################
#######################################################PYGAD-RELATED FUNCTIONS############################################################
##########################################################################################################################################
##########################################################################################################################################

#--------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------instantiate_turbine() METHOD--------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def instantiate_turbine(constargs, solution):
    '''TODO: DOCUMENTATION'''
        
    #--------------------------------------------------------------------------------------------------------------------#
    #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#        
        
    #####
    ##### SETTING CONSTARGS
    #####
    N = None
    omega = None
    rcas = None
    airfoils = None
    mode = None
    for key in constargs.keys():
        if key == "N":
            N = constargs[key]
        elif key == "omega":
            omega = constargs[key]
        elif key == "rcas":
            rcas = constargs[key]
        elif key == "Mode":
            mode = constargs[key]
            assert mode in ["mono", "counter", "pitchopt"], "Provide a valid mode ('mono', 'counter', 'pitchopt'). Current value is {mode}".format(mode=mode)
            
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#            
    
    #####
    ##### RETRIEVING SET OF DESIGN PARAMETERS (PARTIAL PHENOTYPE) FROM SOLUTION (GENOTYPE)
    #####
    Z = int(solution[0])
    sigmahub = solution[1]
    sigmatip = solution[2]
    hub_to_tip_ratio = solution[3]    
    
    #####
    ##### INSTANTIATING MONOPLANE WELLS TURBINE WITH RETRIEVED PHENOTYPE VALUES
    #####
    turb = turbine()   
    turb.create_or_add_stage(omega=omega,
                             N=N,
                             rcas=rcas,
                             hub_to_tip_ratio=hub_to_tip_ratio,
                             Z=Z)
    if mode == 'counter':
        turb.create_or_add_stage(omega=-omega,
                                 N=N,
                                 rcas=rcas,
                                 hub_to_tip_ratio=hub_to_tip_ratio,
                                 Z=Z)
    
    #####
    ##### SOLIDITY-BASED LINEAR CHORD DISTRIBUTION ALONG THE RADIUS
    #####
    # Retrieving radial array for solidity-based radial chord distribution calculation.
    r = turb.turbstage1.gp.r
    # Retrieving tip radial coordinate.
    rtip = r[-1]
    # Computing the parameters for the linear chord distribution along the radius (c0 is the ordinate, m is the slope).
    c0 = 2*np.pi*hub_to_tip_ratio*rtip*(sigmatip - sigmahub)/(Z*(hub_to_tip_ratio - 1))
    m = 2*np.pi*(hub_to_tip_ratio*sigmahub - sigmatip)/(Z*(hub_to_tip_ratio - 1))
    # Computing the linear chord distribution along the radius.
    chord = c0 + m*r
    # Computing the tip percent assuming a constant, physical 2mm-gap.
    tip_percent = (2e-3/chord[-1])*100
    
    #####
    ##### SETTING TURBINE'S RADIAL CHORD AND AIRFOIL DISTRIBUTIONS
    #####
    turb.turbstage1.gp.chord = chord
    turb.turbstage1.gp.tip_percent = tip_percent
    if mode == 'counter':
        turb.turbstage2.gp.chord = chord
        turb.turbstage2.gp.tip_percent = tip_percent
        
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#        
    
    # Return statement.
    return turb

#--------------------------------------------------------------------------------------------------------------------#
#------------------------------------------------fitness_func() METHOD-----------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def fitness_func(ga_instance=None, constargs=None, turb=None, solution=None, solution_idx=None, mode_fitness='NoGA', out='fitness'):    
    '''TODO: DOCUMENTATION'''
    
    
    #-------------------------------------------------------ASSERTIONS--------------------------------------------------#      
    #-------------------------------------------------------ASSERTIONS--------------------------------------------------#      
    #-------------------------------------------------------ASSERTIONS--------------------------------------------------#      
    
    assert not (constargs is None), "Provide a valid 'constargs' parameter."
    assert mode_fitness in ['GA', 'NoGA'], "Provide a valid 'mode_fitness' parameter, either 'GA' or 'NoGA'."
    assert out in ['fitness', 'stfitness'], "Provide a valid 'out' parameter, either 'fitness' or 'stfitness'."
    if mode_fitness == 'GA':
        assert not (ga_instance is None), "Provide a valid 'ga_instance' parameter."
        assert not (solution is None), "Provide a valid 'solution' parameter."
        assert not (solution_idx is None), "Provide a valid 'solution_idx' parameter."
    else:
        assert not (turb is None), "Provide a valid 'turb' parameter."
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#      
    
    #####
    ##### SETTING CONSTARGS
    #####
    polars = None
    cpobj = None
    devobj = None
    Nmin = None
    Nmax = None
    mode = None
    args = ["polars", "cpobjs", "devobjs", "weights", "Nmin", "Nmax", "Mode"]
    for key in constargs.keys():
        if key == "polars":
            polars = constargs[key]
        elif key == "cpobjs":
            cpobjs = constargs[key]
        elif key == "devobjs":
            devobjs = constargs[key]
        elif key == "weights":
            weights = constargs[key]
        elif key == "Nmin":
            Nmin = constargs[key]
        elif key == "Nmax":
            Nmax = constargs[key]
        elif key == "Mode":
            mode = constargs[key]

    #####
    ##### TURBINE INSTANTIATION AND CHECK FOR ALREADY COMPUTED GENES
    #####
    if mode_fitness == 'GA':
        turb = instantiate_turbine(constargs, solution)        
        # Converting solution to tuple-form for serving as dictionary entry.
        tup = tuple(solution)
        if tup in ga_instance.processed_genes.keys():
            # If so, returning fitness value.
            return ga_instance.processed_genes[tup]        
            
    #####
    ##### GENERAL VARIABLE DECLARATION
    #####
    # The 'Phitot', 'Psitot' and 'Pitot' variables are for storing the dimensionless flow-parameter, pressure and power
    # resulting from the succesive application of the BEM method.
    Phitot = list()
    Psitot = list()
    Pitot = list()
    # The 'vxlist' variable stores the input velocities to be BEM-computed.
    vxlist = list(np.arange(1, 46))
    
    #####
    ##### MONO BEM CALCULATIONS
    #####    
    # Conditional for when the algorithm is run in 'mono' mode (monoplane Wells turbine).
    if mode == "mono":
        # Applying BEM method to current phenotype-turbine when running in 'mono' mode.
        turb.BEM(vxlist, polars=polars, inputparam='vx', mode='3DAD')
        # Storing dimensionless variables in previously defined parameters.
        Phitot = turb.Phi
        Psitot = turb.Psi
        Pitot = turb.Pi

    #####
    ##### COUNTER BEM CALCULATIONS
    #####
    # Conditional for when the algorithm is run in 'counter' mode (counterrotating Wells turbine).
    elif mode == "counter":
        #####
        ##### CREATING AUXILIARY VARIABLES AND RUNNING BEMS FOR DIFFERENT ANGULAR PITCH VALUES
        #####
        # Creating auxiliary variables of the dimensionless flow-parameter and pressure for each turbine stage.
        Phi1 = list()
        Psi1 = list()
        Phi2 = list()
        Psi2 = list()
        # For loop over the input velocities to which the BEM method is applied succesively.
        for vx in vxlist:
            turb.BEM([vx], polars=polars, inputparam='vx', mode='3DAD')
            # Storing each turbine stage's dimensionless parameters in the auxiliary variables.
            Phi1.append(turb.turbstage1.it.Phi)
            Psi1.append(turb.turbstage1.eout.Psi)
            Phi2.append(turb.turbstage2.it.Phi)
            Psi2.append(turb.turbstage2.eout.Psi)
            # Storing global (turbine-wise) dimensionless variables in previously defined parameters.
            Phitot.append(turb.Phi[0])
            Psitot.append(turb.Psi[0])
            Pitot.append(turb.Pi[0])
    
    #####
    ##### PITCHOPT BEM CALCULATIONS
    #####
    # Conditional for when the algorithm is run in 'pitchopt' mode (optimal pitch configuration for a monoplane Wells turbine).
    elif mode == "pitchopt":
        # Calling "pitch_optimize" function.
        Phitot, optpowers, optangpitchspow, optpsispow, optpispow, opteffspow = pitch_optimization(turb, polars=polars, vxlist=vxlist)
         
    #####
    ##### MONO ROT-SPEED CALCULATIONS
    #####
    # Conditional for 'mono' case. The rotational speed is computed directly.
    if mode=='mono':
        # Least-square fitting and computation of average slope and ordinate-origin of the phi-psi functional relation.
        s0ave, y0ave = psi_phi_slope_fitter(Phi=turb.Phi, Psi=turb.Psi, interpfit=True)
        # K computation (inverse of s0ave).
        K = [s0ave**(-1)]
        # Computation of the rotational speeds for each cpobj.
        Ns = [(2*s0ave*cpobj*turb.turbstage1.gp.rtip/turb.turbstage1.fp.rho)*60/(2*np.pi) for cpobj in cpobjs]

    #####
    ##### COUNTER ROT-SPEED CALCULATIONS
    #####
    # Conditional for 'counter' case, which owns a double turbine stage. The rotational speed is computed by iteration and assuming
    # that both stages are solidary (rotate at the same speed).
    elif mode=='counter':
        # Least-square fitting and computation of average slope and ordinate-origin of the phi-psi functional relations of both stages.
        s0ave1, y0ave1 = psi_phi_slope_fitter(Phi=Phi1, Psi=Psi1, interpfit=True)
        s0ave2, y0ave2 = psi_phi_slope_fitter(Phi=Phi2, Psi=Psi2, interpfit=True)
        # Least-square fitting and computation of average slope and ordinate-origin of the global phi-psi functional relation.
        s0avetot, y0avetot = psi_phi_slope_fitter(Phi=Phitot, Psi=Psitot, interpfit=True)
        # K computation (inverse of s0avetot).
        K = [s0avetot**(-1)]
        # Computation of psi-standard-deviation based on the assumption of linearity between stds.
        devobjs = [devobj/np.average([s0ave1, s0ave2]) for devobj in devobjs]                       
        # Setting the 'Ns', 'cpobj1s' and 'cpobj2s' lists that will store the values of the computed rotational speeds and the distributed damping factors of the stages.
        Ns = list()
        cpobj1s = list()
        cpobj2s = list()
        for cpobj in cpobjs:
            # Boolean 'N1N2match' variable for checking the convergence of the will-be-computed rotational speeds of the turbine stages.
            N1N2match = False
            # Setting the 'factor' variable to 0.5. It determines how the damping coefficient is distributed between the stages.
            factor = 0.5                
            # While loop until matching the rotational speeds of both turbine stages by modifying the distributive factor of the damping coefficient between stages.
            while not N1N2match:
                # Computing the complementary damping factors of the stages based on the value of the 'factor' variable.
                cpobj1 = cpobj*factor
                cpobj2 = cpobj*(1 - factor)
                # Computing the rotational speeds of the stages.
                N1 = np.abs((2*s0ave1*cpobj1*turb.turbstage1.gp.rtip/turb.turbstage1.fp.rho)*60/(2*np.pi))
                N2 = np.abs((2*s0ave2*cpobj2*turb.turbstage1.gp.rtip/turb.turbstage1.fp.rho)*60/(2*np.pi))
                # Checking for rotational-speed-convergence.
                N1N2match = np.abs(N1 - N2) < 5
                if N1N2match:
                    break
                # Computing the complementary damping factors of the stages based on the value of the 'factor' variable.
                cpobj1 = cpobj*(1 - factor)
                cpobj2 = cpobj*factor
                # Computing the rotational speeds of the stages.
                N1 = np.abs((2*s0ave1*cpobj1*turb.turbstage1.gp.rtip/turb.turbstage1.fp.rho)*60/(2*np.pi))
                N2 = np.abs((2*s0ave2*cpobj2*turb.turbstage1.gp.rtip/turb.turbstage1.fp.rho)*60/(2*np.pi))
                # Checking for rotational-speed-convergence.
                N1N2match = np.abs(N1 - N2) < 5
                # Reducing distributive factor.
                factor -= 0.0001
                # Checking whether the factor has dropped below 0; if so, then break the loop (it has no physical rationale anymore).
                if factor < 0:
                    break
            # Computing average rotational speed (having converged, picking either of the rotational speeds of the turbine stages or the average itself should not have further relevance).
            Ns.append(0.5*(N1 + N2))
            cpobj1s.append(cpobj1)
            cpobj2s.append(cpobj2)

    #####
    ##### PITCHOPT ROT-SPEED CALCULATIONS
    #####
    elif mode=='pitchopt':
        # Least-square fitting and computation of average slope and ordinate-origin of the phi-psi functional relation.            
        s0ave, y0ave = psi_phi_slope_fitter(Phi=Phitot, Psi=optpsispow, interpfit=True)
        # K computation (inverse of s0ave).
        K = [s0ave**(-1)]            
        # Computation of the rotational speeds for each cpobj.
        Ns = [(2*s0ave*cpobj*turb.turbstage1.gp.rtip/turb.turbstage1.fp.rho)*60/(2*np.pi) for cpobj in cpobjs]        
        #####
        ##### RECOMPUTING PITCH OPTIMIZATION
        #####
        # Declaring auxiliary variables for storing optimum dimensionless variables for each recomputed power-optimized sea-state curve.
        phis = list()        
        optpows = list()
        optangs = list()
        optpsis = list()
        optpis = list()
        # Loop running over the calculated rotational speeds for each sea-state.
        for N in Ns:            
        # Setting rotational speed of turbine to calculated output value.
            turb.turbstage1.omega = N*2*np.pi/60
            # Calling "pitch_optimize" function.
            Phitot, optpowers, optangpitchspow, optpsispow, optpispow, opteffspow = pitch_optimization(turb, polars=polars, vxlist=vxlist)
            # Appending results to auxiliary variables.
            phis.append(Phitot)
            optpows.append(optpowers)
            optangs.append(optangpitchspow)
            optpsis.append(optpsispow)
            optpis.append(optpispow)       
            
    #####
    ##### DEVOBJS RECASTING
    #####   
    # Recasting the units of the target standard deviation of the flow parameter.
    devobjs = [devobj/(turb.turbstage1.fp.rho*((Ns[e]*2*np.pi/60)**2)*(2*turb.turbstage1.gp.rtip)**2) for e, devobj in enumerate(devobjs)]   

    #####
    ##### MONO AND COUNTER STOCHASTIC CALCULATIONS AND FITNESS DETERMINATION
    #####
    dev_mininds = list()
    eff_mininds = list()
    fitnesses = list()
    
    #####
    ##### STOCHASTIC CALCULATIONS
    #####
    # Perform the stochastic analysis on the global dimensionless variables of the turbine.    
    if mode=='mono' or mode=='counter':
        # Conditional for when the algorithm is run in either 'mono' or 'counter' mode.
        turb_stoch = stochastic_analysis(Phitot, Psitot, Pitot)        
        
        #####
        ##### FITNESS DETERMINATION
        #####
        # Set the output standard deviations and stochastic efficiencies to 'devs' and 'effs', respectively.
        devs = turb_stoch[0]
        effs = turb_stoch[-1]
        # Interpolate between 'devs' and 'effs' slinearly.
        f = interp1d(devs, effs, kind='slinear', fill_value='extrapolate')
        # Perform the fitting for getting the stochastic efficiency value at the target standard deviation.
        dev_fit = np.linspace(devs[0], devs[-1], 1000)
        eff_fit = f(dev_fit)
        # Initializing fitness value to 0.
        fitness = 0        
        # Loop running over the calculated rotational speeds for the sea-states.
        for e, N in enumerate(Ns):
            # Conditional for checking whether the computed rotational speed stands between the limits of the provided rotational speeds.
            if Nmin < N < Nmax:
                # Compute the index at which the minimum difference is obtained between the target standard deviation and the stochastic
                # efficiency points computed from the fitting.
                minind = np.argmin(np.array([np.abs(_ - devobjs[e]) for _ in dev_fit]))
                dev_mininds.append(dev_fit[minind])
                eff_mininds.append(eff_fit[minind])
                # Summing to the computation of the fitness value.
                fitness += weights[e]*eff_fit[minind]
                fitnesses.append(weights[e]*eff_fit[minind])
                # Conditional for when the algorithm is run in 'counter' mode.
                if mode=='counter':
                    # The 'fitness' values are set to the stochastic efficiency value, the computed rotational speed, and the distributed
                    # damping values between the stages.
                    if eff_fit[minind] == 0:
                        Ns[e] = 0
                        cpobj1s[e] = 0
                        cpobj2s[e] = 0
                    else:
                        pass
            # In case the computed rotational speed does not fall between the limits of the provided rotational speed values, then return
            # a null fitness value, with its corresponding side-values being set to 0 as well.
            else:
                dev_mininds.append([0])
                eff_mininds.append([0])                
                fitness += 0
                fitnesses.append([0])
                Ns[e] = 0
                # Adding null 'cpobj1s' and 'cpobj2s' values for the 'counter' case.
                if mode=='counter':
                    cpobj1s[e] = 0
                    cpobj2s[e] = 0    

    #####
    ##### PITCHOPT STOCHASTIC CALCULATIONS AND FITNESS DETERMINATION
    #####
    elif mode=='pitchopt':
        dev_mins = list()
        eff_mins = list()
        fits = list()
        dev_fit = list()
        eff_fit = list()        
        # Declaring 'vxlist_optangpitchs' variable for storing the optimum pitch angles for each input velocity at each sea-state.
        vxlist_optangpitchs = list()
        # Initializing fitness value to 0.
        fitness = 0        
        # Loop running over the sea-states and computing the stochastic calculations accordingly; storing the values in 'turb_stochs'.
        for e, N in enumerate(Ns):
            
            #####
            ##### STOCHASTIC CALCULATIONS
            #####
            # Conditional for when the algorithm is run in 'pitchopt' mode.            
            turb_stoch = stochastic_analysis(phis[e], optpsis[e], optpis[e])
            
            #####
            ##### FITNESS DETERMINATION
            #####
            # Set the output standard deviations and stochastic efficiencies to 'devs' and 'effs', respectively.
            devs = turb_stoch[0]
            effs = turb_stoch[-1]
            # Interpolate between 'devs' and 'effs' slinearly.
            f = interp1d(devs, effs, kind='slinear', fill_value='extrapolate')
            # Perform the fitting for getting the stochastic efficiency value at the target standard deviation.
            dev_f = np.linspace(devs[0], devs[-1], 1000)            
            eff_f = f(dev_f)            
            # Conditional for checking whether the computed rotational speed stands between the limits of the provided rotational speeds.
            if Nmin < N < Nmax:
                # Compute the index at which the minimum difference is obtained between the target standard deviation and the stochastic
                # efficiency points computed from the fitting.
                dev_fit.append(dev_f)
                eff_fit.append(eff_f)
                minind = np.argmin(np.array([np.abs(_ - devobjs[e]) for _ in dev_f]))
                dev_mins.append(dev_f[minind])
                eff_mins.append(eff_f[minind])                
                # Summing to the computation of the fitness value.
                fitness += weights[e]*eff_f[minind]
                fits.append(weights[e]*eff_f[minind])
                vxlist_optangpitchs.append(list(itertools.chain.from_iterable(zip(vxlist, optangs[e]))))
            # In case the computed rotational speed does not fall between the limits of the provided rotational speed values, then return
            # a null fitness value, with its corresponding side-values being set to 0 as well.
            else:
                dev_fit.append(np.array([0]))
                eff_fit.append(np.array([0]))
                dev_mins.append(0)
                eff_mins.append(0)                                
                fitness += 0                
                fits.append(0)
                Ns[e] = 0
                vxlistaux = [0 for _ in vxlist]
                optangpitchspowaux = [0 for _ in optangs[e]]
                vxlist_optangpitchs.append(list(itertools.chain.from_iterable(zip(vxlistaux, optangpitchspowaux))))
        dev_mininds.append(dev_mins)
        eff_mininds.append(eff_mins)
        fitnesses.append(fits)
    
     # Conditional for instantiating the output 'fitness_out' variable.
    if mode=='mono':
        # In the 'mono' case, the output variable comprises the fitness value, the slope 'K' and the rotational speed 'Ns'.
        fitness_out = [fitness] + K + Ns
        # Recasting the output variable to tuple.
        fitness_out = tuple(fitness_out)
    elif mode=='counter':
        # In the 'counter' case, the output variable comprises the fitness value, the slope 'K' and the set of rotational speeds and cpobj distributions between the turbine stages for each sea-state.
        fitness_out = [fitness] + K
        for e, _ in enumerate(Ns):
            fitness_out += [Ns[e]]
            fitness_out += [cpobj1s[e]]
            fitness_out += [cpobj2s[e]]
        # Recasting the output variable to tuple.
        fitness_out = tuple(fitness_out)
    elif mode=='pitchopt':
        # In the 'pitchopt' case, the output variable comprises the fitness value, the slope 'K', the rotational speed 'Ns' and the set of optimum angular pitches for the input velocities.
        fitness_out = [fitness] + K
        for e, N in enumerate(Ns):
            fitness_out += [Ns[e]]
            fitness_out += vxlist_optangpitchs[e]
        # Recasting the output variable to tuple.
        fitness_out = tuple(fitness_out)
        
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#         
            
    # Return statement.
    if out == 'fitness':
        return fitness_out
    else:
        return dev_fit, eff_fit, dev_mininds, eff_mininds, fitnesses

#--------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------------on_start() METHOD-------------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def on_start(ga_instance, par_to_fit, writepops, rootfilename, pars_to_write, mode):
    '''
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    ##### 
    # In this case, it is used to set a sort of attributes in the ga_instance class
    # that are required in the user-set computation of the algorithm we are using.
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    #####'''

    #--------------------------------------------------------------------------------------------------------------------#
    #-----------------------------------------------------ASSERTIONS-----------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------# 
    
    assert par_to_fit in ["maxeff", "maxstocheff", "maxeffarea", "IDOM"], "Please provide a valid parameter to fit among ['maxeff', 'maxstocheff', 'maxeffarea', 'IDOM']. Current value is " + str(par_to_fit)
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#     
    
    # Setting "processed_genes" attribute as an empty dictionary; it will store the set of all computed genes during the calculation.
    setattr(ga_instance, "processed_genes", dict())
    # Setting "par_to_fit" attribute to "par_to_fit" input variable; it will store the string that tells the algorithm which function to employ when performing the fitting.
    setattr(ga_instance, "par_to_fit", par_to_fit)
    # Conditional for checking wheter the populations and computed genes are to be written to files.
    if writepops:
        # If so, checking whether the 'rootfilename.txt' generic file exists.
        rootfilenameaux = ""
        if os.path.isfile(".".join([rootfilename, "txt"])):
            # If so, then iterate over programatically modified filenames (modified by adding _1, _2, _3...to the root file).
            isfile = True
            i = 1
            # While loop for the modification of the filename.
            while isfile:
                rootfilenameaux = "_".join([_ for _ in rootfilename.split("_")[:-1]] +  [str(i)])
                isfile = os.path.isfile(".".join([rootfilenameaux, "txt"]))
                i += 1
        # Setting 'filename' attribute to the provided/computed 'rootfilename.txt' string. It will store generation-wise populations.
        filename = rootfilenameaux if rootfilenameaux != "" else rootfilename
        setattr(ga_instance, "filename", ".".join([filename, "txt"]))
        pars_to_write_ = [_ for _ in pars_to_write]
        if mode=='mono':
            pars_to_write_ += ["N_sea-state_" + str(i) for i in range(1, 11)]
        elif mode=='counter':
            for i in range(1, 11):
                pars_to_write_.append("N_sea-state_" + str(i))
                pars_to_write_.append("cpobj1_sea-state_" + str(i))
                pars_to_write_.append("cpobj2_sea-state_" + str(i))
        elif mode=='pitchopt':
            for i in range(1, 11):
                pars_to_write_ += ["N_sea_state_" + str(i)]
                for j in range(1, 46):
                    pars_to_write_ += ["vx_" + str(j) + "_N_" + str(i), "pitchopt_" + str(j) + "_N_" + str(i)]
        pars_to_write_str = "        ".join(pars_to_write_) + "\n"
        with open(ga_instance.filename, "w") as f:
            f.write(pars_to_write_str)
            f.write("Generation #1" + "-"*(len("".join(pars_to_write_)) + 8*(len(pars_to_write_) - 1) - len("Generation #1")) + "\n")            
        # Setting 'filename_proc_genes' to the provided/computed 'rootfilename_proc_genes.txt' string. It will store the set of all computed genes during the calculation.
        setattr(ga_instance, "filename_proc_genes", ".".join([filename + "_proc_genes", "txt"]))
    else:
        # Otherwise, setting both 'filename' and 'filename_proc_genes' to mpty strings.
        setattr(ga_instance, "filename", "")
        setattr(ga_instance, "filename_proc_genes", "")    
            
    
#--------------------------------------------------------------------------------------------------------------------#
#-------------------------------------------------on_fitness() METHOD------------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def on_fitness(ga_instance, last_population_fitness):
    '''
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    ##### 
    # In this case, it is used to set a sort of attributes in the ga_instance class
    # that are required in the user-set computation of the algorithm we are using.
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    #####'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Loop over the genes of the currently run population to store them in the 'ga_instance.processed_genes' dictionary.
    for e, _ in enumerate(ga_instance.population):
        tup = tuple(_)
        # If the current gene has already been processed, then pass.
        if tup in ga_instance.processed_genes.keys():
            pass
        # Otherwise, add it to the processed genes' pool.
        else:
            ga_instance.processed_genes[tup] = ga_instance.last_generation_fitness[e]
    
#--------------------------------------------------------------------------------------------------------------------#
#------------------------------------------------on_generation() METHOD----------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def on_generation(ga_instance, pars_to_write, mode):
    '''
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    ##### 
    # In this case, it is used to set a sort of attributes in the ga_instance class
    # that are required in the user-set computation of the algorithm we are using.
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    #####'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#    
    
    # Monitoring (re)printing.
    print('\t Completed generations: ', ga_instance.generations_completed, end='\r')
    # If the provided filename is an empty string (no writing operations have been required) then pass.
    if ga_instance.filename == "":
        pass
    # Otherwise, proceed with the writing.
    else:
        # Performing the file opening and writing operations.
        with open(ga_instance.filename, "a") as f:
            # Instantiating an empty list for storing the data to write.
            write_list = list()            
            if type(ga_instance.last_generation_fitness[0])==np.ndarray:
                order_ind = -len(ga_instance.last_generation_fitness[0])
            else:
                order_ind = -1
            # Loop over the genes of the population for retrieving the genome and its fitness value.
            for e, pop in enumerate(ga_instance.population):
                write_list.append([_ for _ in pop] + [_ for _ in ga_instance.last_generation_fitness[e]])
            # Sorting list data according to fitness value.
            write_list = sorted(write_list, key=lambda x: x[order_ind], reverse=True)
            # Recasting to array.
            write_array = np.array(write_list)
            write_list_str = [str(_) for _ in write_list]            
            fmt = list()
            pars_to_write_ = [_ for _ in pars_to_write]
            if mode=='mono':
                pars_to_write_ += ["N_sea-state_" + str(i) for i in range(1, 11)]
                decimals_list = [0, 3, 3, 3, 3, 3] + 10*[0]
            elif mode=='counter':
                for i in range(1, 11):
                    pars_to_write_.append("N_sea-state_" + str(i))
                    pars_to_write_.append("cpobj1_sea-state_" + str(i))
                    pars_to_write_.append("cpobj2_sea-state_" + str(i))
                decimals_list = [0, 3, 3, 3, 3, 3] + 30*[0]
            elif mode=='pitchopt':
                decimals_list = [0, 3, 3, 3, 3, 3]
                for i in range(1, 11):
                    pars_to_write_ += ["N_sea_state_" + str(i)]
                    decimals_list += [0]
                    for j in range(1, 46):
                        pars_to_write_ += ["vx_" + str(j) + "_N_" + str(i), "pitchopt_" + str(j) + "_N_" + str(i)]
                        decimals_list += [0, 0]
            # Writing to file.
            for _ in write_array:
                fmt = list()
                for g, __ in enumerate(_):
                    if decimals_list[g] == 0:
                        _[g] = int(_[g])
                        diff_length = len(pars_to_write_[g]) - len(str(_[g])[:-2])
                    else:
                        diff_length = len(pars_to_write_[g]) - decimals_list[g] - 2
                    fmt.append("%." + str(decimals_list[g]) + "f" + diff_length*" " + 8*" ")
                np.savetxt(f, np.c_[[_]], fmt="".join(fmt))
            pars_to_write_ = "        ".join(pars_to_write_) + "\n"
            f.write("-"*(len(pars_to_write_)) + "\n")
            if ga_instance.generations_completed == ga_instance.num_generations:
                pass
            else:
                f.write(pars_to_write_)
                f.write("Generation #" + str(ga_instance.generations_completed + 1) + "-"*(len(pars_to_write_) - 1 - len("Generation #" + str(ga_instance.generations_completed + 1))) + "\n")
            # Closing the file.
            f.close()
            
#--------------------------------------------------------------------------------------------------------------------#
#---------------------------------------------------on_stop() METHOD-------------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def on_stop(ga_instance, last_generation_fitness, pars_to_write, mode):
    '''
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    ##### 
    # In this case, it is used to set a sort of attributes in the ga_instance class
    # that are required in the user-set computation of the algorithm we are using.
    #####
    ##### TO-BE DOCUMENTATION/WARNING; ALTHOUGH IT IS A BEST-PRACTICE WAY OF MODIFYING A CALLBACK FUNCTION, THE PROGRAMMER WOULD
    ##### PROBABLY HAVE NOT APPROVED THE USAGE OF THIS CALLBACK FUNCTION AS SUCH.
    ##### IT MODIFIES STANDARD STRUCTURES OF EXTERNALLY-PROVIDED CLASSES.
    #####'''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#     
    
    # If the provided filename is an empty string (no writing operations have been required) then pass.
    if ga_instance.filename_proc_genes == "":
        pass
   # Otherwise, proceed with the writing. 
    else:
        # Performing the file opening and writing operations.
        with open(ga_instance.filename_proc_genes, "w") as f:
            # Instantiating an empty list for storing the data to write.
            write_list = list()
            if type(ga_instance.last_generation_fitness[0])==np.ndarray:
                order_ind = -len(ga_instance.last_generation_fitness[0])
            else:
                order_ind = -1            
            # Loop over the genes of the population for retrieving the genome and its fitness value.
            for _ in ga_instance.processed_genes.keys():
                write_list.append(list(_) + [__ for __ in ga_instance.processed_genes[_]])
            # Sorting list data according to fitness value.
            write_list = sorted(write_list, key=lambda x: x[order_ind], reverse=True)
            # Recasting to array.
            write_array = np.array(write_list)
            # Writing to file.
            write_list_str = [str(_) for _ in write_list]            
            fmt = list()
            pars_to_write_ = [_ for _ in pars_to_write]
            if mode=='mono':
                pars_to_write_ += ["N_sea-state_" + str(i) for i in range(1, 11)]
                decimals_list = [0, 3, 3, 3, 3, 3] + 10*[0]
            elif mode=='counter':
                for i in range(1, 11):
                    pars_to_write_.append("N_sea-state_" + str(i))
                    pars_to_write_.append("cpobj1_sea-state_" + str(i))
                    pars_to_write_.append("cpobj2_sea-state_" + str(i))
                decimals_list = [0, 3, 3, 3, 3, 3] + 30*[0]
            elif mode=='pitchopt':
                decimals_list = [0, 3, 3, 3, 3, 3]
                for i in range(1, 11):
                    pars_to_write_ += ["N_sea_state_" + str(i)]
                    decimals_list += [0]
                    for j in range(1, 46):
                        pars_to_write_ += ["vx_" + str(j) + "_N_" + str(i), "pitchopt_" + str(j) + "_N_" + str(i)]
                        decimals_list += [0, 0]
            # Writing to file.
            pars_to_write_aux = "        ".join(pars_to_write_)
            f.write(pars_to_write_aux + "\n")
            f.write("-"*(len(pars_to_write_aux)) + "\n")
            for _ in write_array:
                fmt = list()
                for g, __ in enumerate(_):
                    if decimals_list[g] == 0:
                        _[g] = int(_[g])
                        diff_length = len(pars_to_write_[g]) - len(str(_[g])[:-2])
                    else:
                        diff_length = len(pars_to_write_[g]) - decimals_list[g] - 2
                    fmt.append("%." + str(decimals_list[g]) + "f" + diff_length*" " + 8*" ")
                np.savetxt(f, np.c_[[_]], fmt="".join(fmt))
            f.write("-"*(len(pars_to_write_aux)) + "\n")
            # Closing the file.
            f.close()

#--------------------------------------------------------------------------------------------------------------------#
#--------------------------------------------cal_pop_single_fitness() METHOD-----------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def cal_pop_single_fitness(ga_instance: pygad.GA,
                           constargs: dict,
                           gene: list,
                           gene_idx: int) -> tuple:
    '''Calculates fitness function for a given gene of the population.
    
    **parameters**, **returns**, **rtypes** ::
    
    :param ga_instance: an instantiated object of the 'pygad.GA' class.
    :param constargs: a dictionary containing the constant arguments required for instantiating a turbine.
    :param gene: a list containing the specific gene to be evaluated.
    :param gene_idx: the index of the gene to be evaluated within the population.
    :returns: a tuple containing the result of applying the fitness function to the gene to be evaluated.
    :rtypes: tuple
    
    '''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------# 
    
    return ga_instance.fitness_func(ga_instance, constargs, gene, gene_idx, mode='GA', out='fitness')

#--------------------------------------------------------------------------------------------------------------------#
#-------------------------------------------------plot_turbine() METHOD----------------------------------------------# 
#--------------------------------------------------------------------------------------------------------------------#
def plot_turbine(turb: turbine) -> None:
    '''Plots front-view of a given turbine desing.
    
    **parameters**, **returns**, **rtypes** ::
    
    :param turb: an instantiated object of the 'turbine' class.
    :returns: a tuple containing a figure object and axes object.
    :rtypes: tuple(plt.figure, plt.figure.gca().axes) 
    
    '''
    
    #--------------------------------------------------------------------------------------------------------------------#
    #--------------------------------------------------------BODY--------------------------------------------------------#
    #--------------------------------------------------------------------------------------------------------------------#     
    
    #####
    ##### GLOBAL TURBINE VARIABLES
    #####    
    # Setting blade numbers to 'Z' variable.
    Z = turb.turbstage1.gp.Z
    # Setting 'tip_percent' to tip-percent.
    tip_percent = turb.turbstage1.gp.tip_percent
    # Setting 'rcas' to casing radius.
    rcas = turb.turbstage1.gp.rcas

    #####
    ##### HUB-RELATED VARIABLES
    #####    
    # Setting 'rhub' to hub radius.
    rhub = turb.turbstage1.gp.rhub
    # Setting 'chub' to hub chord.
    chub = turb.turbstage1.gp.chord[0]
    # Setting 'alphahub0' to angular position of first blade at hub. 
    alphahub0 = np.math.asin(chub/(2*rhub))
    # Setting complex 'zchub1' to complex-positions of the initial placings of the blades at hub.
    zchub1 = [np.math.e**(complex(0, _))*complex(-rhub*np.sin(alphahub0), rhub*np.cos(alphahub0)) for _ in [_*2*np.pi/Z for _ in range(1, Z + 1)]]
    # Setting complex 'zchub2' to complex-positions of the final placings of the blades at hub.
    zchub2 = [np.math.e**(complex(0, _))*complex(rhub*np.sin(alphahub0), rhub*np.cos(alphahub0)) for _ in [_*2*np.pi/Z for _ in range(1, Z + 1)]]    
    
    #####
    ##### TIP-RELATED VARIABLES
    #####
    # Setting 'rtip' to tip radius.
    rtip = turb.turbstage1.gp.rtip
    # Setting 'ctip' to tip chord.
    ctip = turb.turbstage1.gp.chord[-1]
    # Setting 'alphatip0' to angular position of first blade at tip.
    alphatip0 = np.math.asin(ctip/(2*rtip))
    # Setting complex 'zctip1' to complex-positions of the initial placings of the blades at tip.
    zctip1 = [np.math.e**(complex(0, _))*complex(-rtip*np.sin(alphatip0), rtip*np.cos(alphatip0)) for _ in [_*2*np.pi/Z for _ in range(1, Z + 1)]]
    # Setting complex 'zctip2' to complex-positions of the final placings of the blades at tip.
    zctip2 = [np.math.e**(complex(0, _))*complex(rtip*np.sin(alphatip0), rtip*np.cos(alphatip0)) for _ in [_*2*np.pi/Z for _ in range(1, Z + 1)]]
    
    # Declaring 'arcs' and 'arcss' lists to store the arcs that represent the blades sections in a frontal view.
    arcs = list()
    arcss = list()
    
    # Loop running over the hub-tip blade portions.
    for zc1 in [zchub1, zctip1]:
        # Setting 'zc2' portion of hub or tip according to 'zc1' (depending on if it is a hub- or tip-related blade portion).
        zc2 = zchub2 if zc1 == zchub1 else zctip2
        # Setting diameter according to 'zc1' (depending on if it is a hub- or tip-related blade portion).
        d = 2*rhub if zc1 == zchub1 else 2*rtip
        # Loop running over the blade portions.
        for e, _ in enumerate(zc1):
            # Getting signs of hub and tip real and imaginary parts.
            sgnreal1 = np.sign(np.real(zc1[e]))
            sgnimag1 = np.sign(np.imag(zc1[e]))
            sgnreal2 = np.sign(np.real(zc2[e]))
            sgnimag2 = np.sign(np.imag(zc2[e]))
            # Conditionals for setting the angular bounds in degrees by computing the 'offset' and 'multisign' variables for hub and tip.
            if sgnreal1 < 0 and sgnimag1 < 0:
                offset1 = 180
                multsign1 = 1
            elif sgnreal1 > 0 and sgnimag1 < 0:
                offset1 = 360
                multsign1 = 1
            elif sgnreal1 < 0 and sgnimag1 > 0:
                offset1 = 180
                multsign1 = 1
            else:
                offset1 = 0
                multsign1 = 1
            if sgnreal2 < 0 and sgnimag2 < 0:
                offset2 = 180
                multsign2 = 1
            elif sgnreal2 > 0 and sgnimag2 < 0:
                offset2 = 360
                multsign2 = 1
            elif sgnreal2 < 0 and sgnimag2 > 0:
                offset2 = 180
                multsign2 = 1
            else:
                offset2 = 0
                multsign2 = 1
            # Computing the 'theta_1' and 'theta_2' angular bounds for setting the arcs.
            theta_1 = offset1 + multsign1*np.arctan(np.imag(zc1[e])/np.real(zc1[e]))*180/np.pi
            theta_2 = offset2 + multsign2*np.arctan(np.imag(zc2[e])/np.real(zc2[e]))*180/np.pi
            # Conditional for checking that 'theta_1' and 'theta_2' do not depart further than the maximum angular spacing, namely 360/Z.
            if theta_1 > theta_2 and theta_1 - theta_2 < 360/Z:
                st = "1"
                theta1 = theta_2
                theta2 = theta_1
            else:
                if theta_2 - theta_1 > 360/Z:
                    theta1 = theta_2
                    theta2 = theta_1
                else:
                    theta1 = theta_1
                    theta2 = theta_2 
            # Appending patch with computed diameters and angular counds to 'arcs' variable.
            arcs.append(mpl.patches.Arc(xy=(0, 0), width=d, height=d,
                                       theta1=theta1,
                                       theta2=theta2, linewidth=2, color='k'))

    #####
    ##### SETTING MAIN PATCHES
    #####
    # Setting hub circumference in front-view.
    hub = plt.Circle((0, 0), rhub, facecolor='k', alpha=0.3, edgecolor='k', linewidth=2)
    # Setting hub edge in front-view.
    hubedge = plt.Circle((0, 0), rhub, facecolor='None', edgecolor='k', linewidth=2)
    # Setting tip circumference in front-view.
    tip = plt.Circle((0, 0), rtip, facecolor='None', edgecolor='k', linestyle='--')
    # Setting casing circumference in front-view.
    cas = plt.Circle((0, 0), rcas, facecolor='None', edgecolor='k', linestyle='--')
        
    #####
    ##### DECLARING FIGURE
    #####
    fig, ax = plt.subplots(figsize=(10, 10))
    
    #####
    ##### ADDING PATCHES
    #####
    # Adding hub patch.
    ax.add_patch(hub)
    # Adding hub edge.
    ax.add_patch(hubedge)
    # Loop running over the blade portions and plotting lines between hubs and tips.
    for e, _ in enumerate(zchub1):
        ax.plot([np.real(zchub1[e]), np.real(zctip1[e])], [np.imag(zchub1[e]), np.imag(zctip1[e])], color='k', linewidth=2)
        ax.plot([np.real(zchub2[e]), np.real(zctip2[e])], [np.imag(zchub2[e]), np.imag(zctip2[e])], color='k', linewidth=2)
    
    # Adding tip patch.
    ax.add_patch(tip)
    # Loop running over the blade portions and adding arc patches.
    for arc in arcs:
        ax.add_patch(arc) 
    
    #####
    ##### FIGURE SPECIFICATIONS
    #####        
    # Setting x-axis limits.
    ax.set_xlim((-rcas*1.1, rcas*1.05))
    # Setting y-axis limits.
    ax.set_ylim((-rcas*1.1, rcas*1.05))
    # Disabling ticks and axes.
    ax.set_xticks([])
    ax.set_yticks([])
    ax.axis('off')    
    # Annotating hub radius.
    ax.arrow(0, 0, 0.95*rhub/np.sqrt(2), 0.95*rhub/np.sqrt(2))
    ax.annotate('$r_{hub}$=' + str(rhub)[:4], xy=(0.75*rhub, 0.75*rhub), fontsize=20)
    # Annotating tip radius.
    ax.arrow(0, 0, -0.99*rtip/np.sqrt(2), 0.95*rtip/np.sqrt(2))
    ax.annotate('$r_{tip}$=' + str(rtip)[:4], xy=(-1.01*rtip, 0.7*rtip), fontsize=20)
    # Annotating blade number.
    ax.annotate('Z={Z}'.format(Z=Z), xy=(-0.15*rhub, -rhub/5), fontsize=20)
    # Annotating hub-to-tip-ratio.
    ax.annotate('$\\nu$={hub_to_tip_ratio:.2f}'.format(hub_to_tip_ratio=rhub/rtip), xy=(-0.25*rhub, -2*rhub/5), fontsize=20)
    # Annotating hub and tip chords.
    ax.annotate('$c_{hub}$=' + str(chub)[:4] + '; $c_{tip}$=' + str(ctip)[:4], xy=(-0.65*rhub, -3*rhub/5), fontsize=20)
    # Annotating tip percent.
    ax.annotate('gap={tip_percent}'.format(tip_percent=np.round(tip_percent, 2)), xy=(-0.25*rhub, -4*rhub/5), fontsize=20)

    #--------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------------RETURN-------------------------------------------------------#             
    #--------------------------------------------------------------------------------------------------------------------#     
    
    # Return statement.
    return fig, ax